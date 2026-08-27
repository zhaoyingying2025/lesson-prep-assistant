"""API 路由汇总（MVP简化版，所有路由聚合便于维护）"""
from __future__ import annotations

import asyncio
import hashlib
import json
import logging
import re
import urllib.parse
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import Response
from sqlalchemy import func, select
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession

from ..agents.chat_agent import modify_lesson
from ..agents.knowledge_agent import (
    chunked_extract_knowledge,
    extract_knowledge,
    organize_knowledge_into_chapters,
)
from ..agents.lesson_agent import generate_lesson
from ..agents.lesson_agent_addie import generate_lesson_with_addie
from ..agents.material_evaluator import evaluate_lesson_independent
from ..agents.ppt_agent import generate_ppt_content
from ..agents.smart_extract_agent import smart_extract
from ..core.llm import (
    LLMError,
    PRESET_PROVIDERS,
    get_llm,
    get_llm_config,
    reset_llm,
)
from ..core.parser import ParseError, detect_material_type, make_preview, parse_file
from ..core.content_validator import validate_lesson, validate_ppt
from ..core.textbook_cache import index_material, search_textbook, get_knowledge_point_context
from ..exporters.docx_export import export_docx
from ..exporters.markdown import export_markdown
from ..exporters.pptx_export import export_teaching_pptx
from ..exporters.template_docx import generate_template_docx, parse_template_docx
from ..models.schemas import (
    ApiResponse,
    ChatRequest,
    CourseCreate,
    LEGACY_MATERIAL_TYPE_MAP,
    LessonParams,
    LessonPlan,
    LessonTemplateCreate,
    LessonTemplateUpdate,
    LLMSettingsUpdate,
)
from ..storage.db import (
    ChatMessageORM,
    ChapterORM,
    CourseORM,
    KnowledgePointORM,
    LessonORM,
    LessonTemplateORM,
    MaterialORM,
    PptRecordORM,
    get_session,
)
from ..storage.file_store import save_upload

# 六类合法教材类型枚举
VALID_MATERIAL_TYPES = {"syllabus", "textbook", "reference", "exercise_book", "paper", "other"}
# 六类中文标签（用于前端展示）
MATERIAL_TYPE_LABELS = {
    "syllabus": "课程标准/大纲",
    "textbook": "教科书",
    "reference": "教参教辅",
    "exercise_book": "练习题册",
    "paper": "学术论文",
    "other": "其他",
}
# 教案/PPT生成失败时的fallback操作集合
FALLBACK_ACTIONS = ["retry", "lower_params", "template_fallback", "export_draft", "open_help"]


def _normalize_material_type(raw: str | None) -> tuple[str, str | None]:
    """规范化教材类型：支持legacy映射、非法值回退other；返回 (type, hint_or_None)"""
    if not raw:
        return ("other", None)
    t = raw.strip()
    # legacy映射
    if t in LEGACY_MATERIAL_TYPE_MAP:
        return (LEGACY_MATERIAL_TYPE_MAP[t], f"类型已兼容：旧值 {t} → {LEGACY_MATERIAL_TYPE_MAP[t]}")
    if t in VALID_MATERIAL_TYPES:
        return (t, None)
    return ("other", f"类型已修正：未知类型 {t} → other")


router = APIRouter(prefix="/api")

# 知识点提取进度跟踪（内存存储，单进程适用）
_extract_progress: dict[int, dict] = {}

async def _update_extract_progress(course_id: int, current: int, total: int, points: list):
    _extract_progress[course_id] = {
        "current": current,
        "total": total,
        "points": points,
        "updated_at": datetime.utcnow().isoformat(),
    }


async def _compute_course_content_hash(session: AsyncSession, course_id: int) -> str:
    """计算课程所有教材内容的 SHA256 哈希值，用于增量提取缓存判断"""
    result = await session.execute(
        select(MaterialORM).where(MaterialORM.course_id == course_id)
    )
    materials = result.scalars().all()
    texts = []
    for m in materials:
        if m.content_text:
            texts.append(m.content_text)
    combined = "\n".join(texts)
    return hashlib.sha256(combined.encode("utf-8")).hexdigest() if combined else ""


# ============================================================
# 课程管理
# ============================================================
@router.get("/courses")
async def list_courses():
    async for session in get_session():
        result = await session.execute(select(CourseORM).order_by(CourseORM.created_at.desc()))
        courses = result.scalars().all()
        return ApiResponse(
            data=[
                {
                    "id": c.id,
                    "name": c.name,
                    "major": c.major,
                    "description": c.description,
                    "subject": c.subject or "",
                    "created_at": c.created_at.isoformat() if c.created_at else None,
                }
                for c in courses
            ]
        )


@router.post("/courses")
async def create_course(payload: CourseCreate):
    async for session in get_session():
        course = CourseORM(
            name=payload.name,
            major=payload.major,
            description=payload.description,
            subject=payload.subject or "",
        )
        session.add(course)
        await session.commit()
        await session.refresh(course)
        return ApiResponse(
            message="课程创建成功",
            data={
                "id": course.id,
                "name": course.name,
                "major": course.major,
                "description": course.description,
                "subject": course.subject or "",
                "created_at": course.created_at.isoformat() if course.created_at else None,
            },
        )


@router.get("/courses/{course_id}")
async def get_course(course_id: int):
    async for session in get_session():
        course = await session.get(CourseORM, course_id)
        if not course:
            raise HTTPException(404, "课程不存在")
        return ApiResponse(
            data={
                "id": course.id,
                "name": course.name,
                "major": course.major,
                "description": course.description,
                "subject": course.subject or "",
                "created_at": course.created_at.isoformat() if course.created_at else None,
            }
        )


@router.delete("/courses/{course_id}")
async def delete_course(course_id: int):
    async for session in get_session():
        course = await session.get(
            CourseORM,
            course_id,
            options=[
                selectinload(CourseORM.materials),
                selectinload(CourseORM.ppt_records),
            ],
        )
        if not course:
            raise HTTPException(404, "课程不存在")

        # 1. 先删除上传的文件（教材、PPT存储文件）
        from ..config import settings
        base_dir = Path(settings.upload_dir) if settings.upload_dir else Path(__file__).resolve().parent.parent.parent / "uploads"

        for mat in course.materials or []:
            if mat.stored_path:
                p = Path(mat.stored_path)
                # 仅删除 uploads 目录下的文件，避免误删
                try:
                    if p.exists() and base_dir in p.resolve().parents:
                        p.unlink()
                except Exception:
                    pass

        for ppt in course.ppt_records or []:
            if ppt.stored_path:
                p = Path(ppt.stored_path)
                try:
                    if p.exists() and base_dir in p.resolve().parents:
                        p.unlink()
                except Exception:
                    pass

        # 2. 使用 SQLAlchemy 级联删除
        await session.delete(course)
        await session.commit()
        return ApiResponse(message="课程已删除")


@router.put("/courses/{course_id}")
async def update_course(course_id: int, payload: dict):
    """更新课程信息（名称/专业/描述/学科）"""
    async for session in get_session():
        course = await session.get(CourseORM, course_id)
        if not course:
            raise HTTPException(404, "课程不存在")
        if "name" in payload:
            name = (payload.get("name") or "").strip()
            if not name:
                raise HTTPException(400, "课程名称不能为空")
            course.name = name
        if "major" in payload:
            course.major = (payload.get("major") or "").strip() or None
        if "description" in payload:
            course.description = (payload.get("description") or "").strip() or None
        if "subject" in payload:
            course.subject = (payload.get("subject") or "").strip()
        await session.commit()
        await session.refresh(course)
        return ApiResponse(
            message="课程已更新",
            data={
                "id": course.id,
                "name": course.name,
                "major": course.major,
                "description": course.description,
                "subject": course.subject or "",
                "created_at": course.created_at.isoformat() if course.created_at else None,
            },
        )


# ============================================================
# 章节树（章/节/子节，多级嵌套）
# ============================================================
def _chapter_to_dict(ch: ChapterORM, include_children: bool = True) -> dict:
    """递归序列化章节树"""
    d = {
        "id": ch.id,
        "course_id": ch.course_id,
        "parent_id": ch.parent_id,
        "name": ch.name,
        "sort_order": ch.sort_order,
    }
    if include_children:
        d["children"] = [_chapter_to_dict(c) for c in (ch.children or [])]
    return d


@router.get("/courses/{course_id}/chapters")
async def list_chapters(course_id: int):
    """获取课程的章节树（递归嵌套）"""
    async for session in get_session():
        # 一次查出该课程的所有章节，在 Python 中构建树（避免异步 lazy load 问题）
        result = await session.execute(
            select(ChapterORM)
            .where(ChapterORM.course_id == course_id)
            .order_by(ChapterORM.sort_order)
        )
        all_chapters = result.scalars().all()

        # 构建 id -> node 字典
        node_map = {}
        for ch in all_chapters:
            node_map[ch.id] = {
                "id": ch.id,
                "course_id": ch.course_id,
                "parent_id": ch.parent_id,
                "name": ch.name,
                "sort_order": ch.sort_order,
                "children": [],
            }

        # 构建树（收集根节点）
        roots = []
        for ch in all_chapters:
            node = node_map[ch.id]
            if ch.parent_id is None:
                roots.append(node)
            else:
                parent = node_map.get(ch.parent_id)
                if parent:
                    parent["children"].append(node)
                else:
                    # 父节点不存在，作为根节点
                    roots.append(node)

        return ApiResponse(data=roots)


@router.post("/courses/{course_id}/chapters")
async def create_chapter(course_id: int, payload: dict):
    """新建章节（parent_id 为 null 表示顶级章，否则为子节）"""
    name = (payload.get("name") or "").strip()
    if not name:
        raise HTTPException(400, "章节名称不能为空")
    parent_id = payload.get("parent_id")

    async for session in get_session():
        # 校验课程存在
        course = await session.get(CourseORM, course_id)
        if not course:
            raise HTTPException(404, "课程不存在")
        # 校验父节点
        if parent_id is not None:
            parent = await session.get(ChapterORM, parent_id)
            if not parent or parent.course_id != course_id:
                raise HTTPException(400, "父节点不存在或不属于本课程")

        # 计算 sort_order：同级末尾
        if parent_id is not None:
            result = await session.execute(
                select(func.max(ChapterORM.sort_order))
                .where(ChapterORM.parent_id == parent_id)
            )
        else:
            result = await session.execute(
                select(func.max(ChapterORM.sort_order))
                .where(ChapterORM.course_id == course_id, ChapterORM.parent_id.is_(None))
            )
        max_order = result.scalar() or 0

        chapter = ChapterORM(
            course_id=course_id,
            parent_id=parent_id,
            name=name,
            sort_order=max_order + 1,
        )
        session.add(chapter)
        await session.commit()
        await session.refresh(chapter)
        # 直接构造 dict，避免访问 relationship 触发异步 lazy load
        return ApiResponse(data={
            "id": chapter.id,
            "course_id": chapter.course_id,
            "parent_id": chapter.parent_id,
            "name": chapter.name,
            "sort_order": chapter.sort_order,
            "children": [],
        }, message="章节已创建")


@router.put("/chapters/{chapter_id}")
async def rename_chapter(chapter_id: int, payload: dict):
    """重命名章节"""
    name = (payload.get("name") or "").strip()
    if not name:
        raise HTTPException(400, "章节名称不能为空")

    async for session in get_session():
        chapter = await session.get(ChapterORM, chapter_id)
        if not chapter:
            raise HTTPException(404, "章节不存在")
        chapter.name = name
        await session.commit()
        await session.refresh(chapter)
        return ApiResponse(data={
            "id": chapter.id,
            "course_id": chapter.course_id,
            "parent_id": chapter.parent_id,
            "name": chapter.name,
            "sort_order": chapter.sort_order,
            "children": [],
        }, message="已重命名")


@router.delete("/chapters/{chapter_id}")
async def delete_chapter(chapter_id: int):
    """删除章节（级联删除子节点；关联教案 chapter_id 置空，保留教案）"""
    async for session in get_session():
        chapter = await session.get(ChapterORM, chapter_id)
        if not chapter:
            raise HTTPException(404, "章节不存在")
        await session.delete(chapter)
        await session.commit()
        return ApiResponse(message="章节已删除")


# ============================================================
# 教材资源
# ============================================================
@router.post("/courses/{course_id}/materials")
async def upload_material(
    course_id: int,
    file: UploadFile = File(...),
    material_type: Optional[str] = Form(None),
    version_label: Optional[str] = Form(None),
    is_primary: Optional[bool] = Form(None),
):
    async for session in get_session():
        course = await session.get(CourseORM, course_id)
        if not course:
            raise HTTPException(404, "课程不存在")

        # 读取文件
        file_bytes = await file.read()
        if not file_bytes:
            raise HTTPException(400, "文件为空")

        # 保存文件
        stored_path = await save_upload(file_bytes, file.filename or "unnamed", course_id)

        # 解析文件
        try:
            content_text = parse_file(Path(stored_path))
        except ParseError as e:
            # 解析失败也要保存记录，但内容为空
            content_text = ""
            parse_error = str(e)
        else:
            parse_error = None

        # 类型检测：优先用户预选 → 启发式检测 → 规范化
        raw_type = material_type or detect_material_type(file.filename or "", content_text)
        detected_type, type_hint = _normalize_material_type(raw_type)

        # 若设为主教材,先清除该课程下其他教材的主教材标记(同课程仅一本主教材)
        primary_flag = bool(is_primary) if is_primary is not None else False
        if primary_flag:
            await session.execute(
                __import__("sqlalchemy").update(MaterialORM)
                .where(MaterialORM.course_id == course_id)
                .values(is_primary=False)
            )

        # 入库
        material = MaterialORM(
            course_id=course_id,
            filename=file.filename or "unnamed",
            stored_path=str(stored_path),
            material_type=detected_type,
            version_label=(version_label or "").strip(),
            is_primary=primary_flag,
            file_size=len(file_bytes),
            content_text=content_text,
            content_preview=make_preview(content_text),
            char_count=len(content_text),
        )
        session.add(material)
        await session.commit()
        await session.refresh(material)

        # 自动索引到教材缓存（按页分割，便于后续检索）
        if content_text:
            try:
                indexed_pages = await index_material(
                    session, material.id, course_id, Path(stored_path)
                )
            except Exception:
                indexed_pages = 0
        else:
            indexed_pages = 0

        # 更新课程内容哈希（增量缓存）
        course.content_hash = await _compute_course_content_hash(session, course_id)
        await session.commit()

        message_parts = ["文件上传成功"]
        if parse_error:
            message_parts.append("（部分内容解析失败）")
        if type_hint:
            message_parts.append(f"（{type_hint}）")
        if primary_flag:
            message_parts.append("（已设为主教材）")
        return ApiResponse(
            message="".join(message_parts),
            data={
                "id": material.id,
                "course_id": material.course_id,
                "filename": material.filename,
                "material_type": material.material_type,
                "material_type_label": MATERIAL_TYPE_LABELS.get(material.material_type, "其他"),
                "version_label": material.version_label or "",
                "is_primary": material.is_primary,
                "file_size": material.file_size,
                "char_count": material.char_count,
                "content_preview": material.content_preview,
                "indexed_pages": indexed_pages,
                "parse_error": parse_error,
                "created_at": material.created_at.isoformat() if material.created_at else None,
            },
        )


@router.put("/materials/{material_id}")
async def update_material(material_id: int, payload: dict):
    """修改教材信息（material_type / filename / version_label / is_primary）"""
    async for session in get_session():
        m = await session.get(MaterialORM, material_id)
        if not m:
            raise HTTPException(404, "材料不存在")
        hint = ""
        new_type = payload.get("material_type")
        if new_type is not None:
            norm_type, hint = _normalize_material_type(new_type)
            m.material_type = norm_type
        new_filename = payload.get("filename")
        if new_filename and isinstance(new_filename, str):
            m.filename = new_filename[:255]
        new_version = payload.get("version_label")
        if new_version is not None and isinstance(new_version, str):
            m.version_label = new_version.strip()[:100]
        # 设置主教材：同课程仅一本主教材
        if "is_primary" in payload:
            primary_flag = bool(payload.get("is_primary"))
            if primary_flag:
                await session.execute(
                    __import__("sqlalchemy").update(MaterialORM)
                    .where(MaterialORM.course_id == m.course_id)
                    .values(is_primary=False)
                )
            m.is_primary = primary_flag
        await session.commit()
        await session.refresh(m)
        msg = "教材信息已更新"
        if hint:
            msg += f"（{hint}）"
        if m.is_primary:
            msg += "（主教材）"
        return ApiResponse(
            message=msg,
            data={
                "id": m.id,
                "filename": m.filename,
                "material_type": m.material_type,
                "material_type_label": MATERIAL_TYPE_LABELS.get(m.material_type, "其他"),
                "version_label": m.version_label or "",
                "is_primary": m.is_primary,
            },
        )


@router.put("/materials/{material_id}/reupload")
async def reupload_material(material_id: int, file: UploadFile = File(...)):
    """重新上传教材文件（替换文件内容，保留原记录ID和元数据）"""
    async for session in get_session():
        m = await session.get(MaterialORM, material_id)
        if not m:
            raise HTTPException(404, "材料不存在")
        file_bytes = await file.read()
        if not file_bytes:
            raise HTTPException(400, "文件为空")
        # 删除旧文件
        try:
            Path(m.stored_path).unlink(missing_ok=True)
        except Exception:
            pass
        # 保存新文件
        stored_path = await save_upload(file_bytes, file.filename or m.filename, m.course_id)
        # 解析新文件
        try:
            content_text = parse_file(Path(stored_path))
        except ParseError:
            content_text = ""
        m.filename = file.filename or m.filename
        m.stored_path = str(stored_path)
        m.file_size = len(file_bytes)
        m.content_text = content_text
        m.content_preview = make_preview(content_text)
        m.char_count = len(content_text)
        await session.commit()

        # 更新课程内容哈希（增量缓存）
        course = await session.get(CourseORM, m.course_id)
        if course:
            course.content_hash = await _compute_course_content_hash(session, m.course_id)
            await session.commit()

        await session.refresh(m)
        return ApiResponse(
            message="教材文件已替换",
            data={
                "id": m.id,
                "filename": m.filename,
                "material_type": m.material_type,
                "file_size": m.file_size,
                "char_count": m.char_count,
                "content_preview": m.content_preview,
            },
        )


@router.get("/courses/{course_id}/materials")
async def list_materials(course_id: int):
    async for session in get_session():
        result = await session.execute(
            select(MaterialORM)
            .where(MaterialORM.course_id == course_id)
            .order_by(MaterialORM.created_at.desc())
        )
        materials = result.scalars().all()
        return ApiResponse(
            data=[
                {
                    "id": m.id,
                    "course_id": m.course_id,
                    "filename": m.filename,
                    "material_type": m.material_type,
                    "material_type_label": MATERIAL_TYPE_LABELS.get(m.material_type, "其他"),
                    "version_label": m.version_label or "",
                    "is_primary": bool(m.is_primary),
                    "file_size": m.file_size,
                    "char_count": m.char_count,
                    "content_preview": m.content_preview,
                    "created_at": m.created_at.isoformat() if m.created_at else None,
                }
                for m in materials
            ]
        )


@router.get("/materials/types")
async def list_material_types():
    """返回六类教材类型枚举：value + label + 颜色，便于前端渲染单选和badge"""
    return ApiResponse(data={
        "types": [
            {"value": "syllabus", "label": "课程标准/大纲", "color": "bg-purple-100 text-purple-700 border-purple-200"},
            {"value": "textbook", "label": "教科书", "color": "bg-emerald-100 text-emerald-700 border-emerald-200"},
            {"value": "reference", "label": "教参教辅", "color": "bg-sky-100 text-sky-700 border-sky-200"},
            {"value": "exercise_book", "label": "练习题册", "color": "bg-amber-100 text-amber-700 border-amber-200"},
            {"value": "paper", "label": "学术论文", "color": "bg-rose-100 text-rose-700 border-rose-200"},
            {"value": "other", "label": "其他", "color": "bg-slate-100 text-slate-700 border-slate-200"},
        ]
    })


@router.put("/materials/{material_id}/set-primary")
async def set_primary_material(material_id: int):
    """将该教材设为课程的主教材（同课程仅一本主教材，AI生成时优先检索）"""
    async for session in get_session():
        m = await session.get(MaterialORM, material_id)
        if not m:
            raise HTTPException(404, "材料不存在")
        # 清除该课程下其他教材的主教材标记
        await session.execute(
            __import__("sqlalchemy").update(MaterialORM)
            .where(MaterialORM.course_id == m.course_id)
            .values(is_primary=False)
        )
        m.is_primary = True
        await session.commit()
        await session.refresh(m)
        return ApiResponse(
            message=f"已将「{m.filename}」设为主教材",
            data={
                "id": m.id,
                "filename": m.filename,
                "is_primary": m.is_primary,
                "version_label": m.version_label or "",
            },
        )


@router.get("/materials/{material_id}")
async def get_material(material_id: int):
    async for session in get_session():
        m = await session.get(MaterialORM, material_id)
        if not m:
            raise HTTPException(404, "材料不存在")
        return ApiResponse(
            data={
                "id": m.id,
                "course_id": m.course_id,
                "filename": m.filename,
                "material_type": m.material_type,
                "file_size": m.file_size,
                "char_count": m.char_count,
                "content_text": m.content_text,
                "content_preview": m.content_preview,
                "created_at": m.created_at.isoformat() if m.created_at else None,
            }
        )


@router.delete("/materials/{material_id}")
async def delete_material(material_id: int):
    async for session in get_session():
        m = await session.get(MaterialORM, material_id)
        if not m:
            raise HTTPException(404, "材料不存在")
        course_id = m.course_id
        # 删除文件
        try:
            Path(m.stored_path).unlink(missing_ok=True)
        except Exception:
            pass
        await session.delete(m)
        await session.commit()
        # 更新课程内容哈希（增量缓存）
        course = await session.get(CourseORM, course_id)
        if course:
            course.content_hash = await _compute_course_content_hash(session, course_id)
            await session.commit()
        return ApiResponse(message="材料已删除")


# ============================================================
# 知识点提取
# ============================================================
@router.post("/courses/{course_id}/extract-knowledge")
async def extract_knowledge_api(
    course_id: int,
    chapter: str = Form(...),
    material_ids: Optional[str] = Form(None),  # 逗号分隔的ID
):
    async for session in get_session():
        course = await session.get(CourseORM, course_id)
        if not course:
            raise HTTPException(404, "课程不存在")

        # 获取材料
        if material_ids:
            ids = [int(x) for x in material_ids.split(",") if x.strip()]
        else:
            result = await session.execute(
                select(MaterialORM).where(MaterialORM.course_id == course_id)
            )
            ids = [m.id for m in result.scalars().all()]

        if not ids:
            raise HTTPException(400, "请先上传教材资源")

        # 拼接材料文本
        texts: list[str] = []
        for mid in ids:
            m = await session.get(MaterialORM, mid)
            if m and m.content_text:
                texts.append(f"### 来源：{m.filename}\n{m.content_text}")

        combined = "\n\n".join(texts)
        if not combined.strip():
            raise HTTPException(400, "教材内容为空，请上传可解析的文件")

        # 增量缓存判断：内容未变化则跳过 LLM 调用，直接返回已有知识点
        current_hash = await _compute_course_content_hash(session, course_id)
        if current_hash and current_hash == course.content_hash:
            cached_kps = (await session.execute(
                select(KnowledgePointORM).where(
                    KnowledgePointORM.course_id == course_id,
                    KnowledgePointORM.chapter == chapter,
                ).order_by(KnowledgePointORM.sort_order)
            )).scalars().all()
            if cached_kps:
                points_with_id = [
                    {
                        "id": kp.id,
                        "name": kp.name,
                        "definition": kp.definition,
                        "source_pages": kp.source_pages or "",
                        "layer": kp.layer,
                        "is_key_point": bool(kp.is_key_point),
                        "is_difficult": bool(kp.is_difficult),
                        "is_exam_point": bool(kp.is_exam_point),
                        "sort_order": kp.sort_order,
                    }
                    for kp in cached_kps
                ]
                return ApiResponse(
                    message=f"教材内容未变化，已从缓存加载 {len(points_with_id)} 个知识点",
                    data={"points": points_with_id, "summary": "（缓存数据，教材内容无变化）"},
                )

        try:
            result = await extract_knowledge(
                course.name, chapter, combined,
                subject=course.subject or None,
            )
        except LLMError as e:
            raise HTTPException(502, str(e))

        # 持久化知识点（先删除该章节旧知识点，再写入新的）
        # 通过 chapter 名匹配（兼容无 chapter_id 的场景）
        old_kps = (await session.execute(
            select(KnowledgePointORM).where(
                KnowledgePointORM.course_id == course_id,
                KnowledgePointORM.chapter == chapter,
            )
        )).scalars().all()
        for old_kp in old_kps:
            await session.delete(old_kp)

        # 写入新知识点
        saved_points = []
        for idx, p in enumerate(result.points, 1):
            kp = KnowledgePointORM(
                course_id=course_id,
                chapter=chapter,
                name=p.name,
                definition=p.definition or "",
                source_pages=p.source_pages or "",
                layer=p.layer if p.layer in ("basic", "core", "extension") else "basic",
                is_key_point=1 if p.is_key_point else 0,
                is_difficult=1 if p.is_difficult else 0,
                is_exam_point=1 if p.is_exam_point else 0,
                sort_order=idx,
            )
            session.add(kp)
            saved_points.append(kp)
        await session.commit()

        # 更新增量缓存标记
        course.content_hash = await _compute_course_content_hash(session, course_id)
        course.last_extracted_at = datetime.utcnow()
        await session.commit()

        for kp in saved_points:
            await session.refresh(kp)

        # 构造返回数据（带 id）
        points_with_id = [
            {
                "id": kp.id,
                "name": kp.name,
                "definition": kp.definition,
                "source_pages": kp.source_pages or "",
                "layer": kp.layer,
                "is_key_point": bool(kp.is_key_point),
                "is_difficult": bool(kp.is_difficult),
                "is_exam_point": bool(kp.is_exam_point),
                "sort_order": kp.sort_order,
            }
            for kp in saved_points
        ]

        return ApiResponse(
            message=f"已提取 {len(points_with_id)} 个知识点",
            data={"points": points_with_id, "summary": result.summary},
        )


# ============================================================
# 智能章节提取（一键知识点提取）
# ============================================================
@router.post("/courses/{course_id}/smart-extract")
async def smart_extract_api(course_id: int):
    """从教材全文自动识别章节结构并提取知识点"""
    async for session in get_session():
        course = await session.get(CourseORM, course_id)
        if not course:
            raise HTTPException(404, "课程不存在")

        # 获取所有教材
        result = await session.execute(
            select(MaterialORM).where(MaterialORM.course_id == course_id)
        )
        materials = result.scalars().all()
        if not materials:
            raise HTTPException(400, "请先上传教材资源")

        # 拼接材料文本
        texts: list[str] = []
        filenames: list[str] = []
        for m in materials:
            if m.content_text:
                texts.append(f"### 来源：{m.filename}\n{m.content_text}")
                filenames.append(m.filename)

        combined = "\n\n".join(texts)
        if not combined.strip():
            raise HTTPException(400, "教材内容为空")

        # 增量缓存判断：内容未变化则跳过 LLM 调用，直接返回已有章节和知识点
        current_hash = await _compute_course_content_hash(session, course_id)
        if current_hash and current_hash == course.content_hash:
            existing_chapters = (await session.execute(
                select(ChapterORM).where(
                    ChapterORM.course_id == course_id,
                    ChapterORM.parent_id.is_(None),
                ).order_by(ChapterORM.sort_order)
            )).scalars().all()
            if existing_chapters:
                kp_count = (await session.execute(
                    select(func.count(KnowledgePointORM.id)).where(
                        KnowledgePointORM.course_id == course_id,
                    )
                )).scalar() or 0
                cached_chapters = []
                async def _build_chapter_node(ch: ChapterORM) -> dict:
                    children = []
                    for child in (ch.children or []):
                        child_node = await _build_chapter_node(child)
                        children.append(child_node)
                    kps = (await session.execute(
                        select(KnowledgePointORM).where(
                            KnowledgePointORM.course_id == course_id,
                            KnowledgePointORM.chapter_id == ch.id,
                        ).order_by(KnowledgePointORM.sort_order)
                    )).scalars().all()
                    return {
                        "id": ch.id,
                        "name": ch.name,
                        "children": children,
                        "knowledge_points": [_kp_to_dict(kp) for kp in kps],
                    }
                for ch in existing_chapters:
                    cached_chapters.append(await _build_chapter_node(ch))
                return ApiResponse(
                    message=f"教材内容未变化，已从缓存加载 {len(cached_chapters)} 个章节，{kp_count} 个知识点",
                    data={"chapters": cached_chapters, "total_kp": kp_count},
                )

        try:
            chapter_tree = await smart_extract(
                course.name,
                ", ".join(filenames),
                combined,
                subject=course.subject or None,
            )
        except LLMError as e:
            raise HTTPException(502, str(e))

        if not chapter_tree:
            raise HTTPException(500, "未能识别出章节结构，请检查教材内容")

        # 递归创建章节和知识点
        created_chapters = []

        async def _create_node(
            parent_id: int | None,
            sort_order: int,
            node: dict,
        ) -> dict:
            ch = ChapterORM(
                course_id=course_id,
                parent_id=parent_id,
                name=node["name"],
                sort_order=sort_order,
            )
            session.add(ch)
            await session.flush()
            await session.refresh(ch)

            node_kps = node.get("knowledge_points") or []
            saved_kps = []
            for kp_idx, kp in enumerate(node_kps, 1):
                kp_obj = KnowledgePointORM(
                    course_id=course_id,
                    chapter=node["name"],
                    chapter_id=ch.id,
                    name=kp["name"],
                    definition=kp.get("definition", ""),
                    source_pages=kp.get("source_pages", ""),
                    layer=kp.get("layer", "core"),
                    is_key_point=1 if kp.get("is_key_point") else 0,
                    is_difficult=1 if kp.get("is_difficult") else 0,
                    is_exam_point=1 if kp.get("is_exam_point") else 0,
                    sort_order=kp_idx,
                )
                session.add(kp_obj)
                saved_kps.append(kp_obj)

            children = []
            for child_idx, child in enumerate(node.get("children") or []):
                child_result = await _create_node(ch.id, child_idx + 1, child)
                children.append(child_result)

            await session.flush()
            for kp in saved_kps:
                await session.refresh(kp)

            return {
                "id": ch.id,
                "name": ch.name,
                "children": children,
                "knowledge_points": [
                    {
                        "id": kp.id,
                        "name": kp.name,
                        "definition": kp.definition,
                        "source_pages": kp.source_pages or "",
                        "layer": kp.layer,
                        "is_key_point": bool(kp.is_key_point),
                        "is_difficult": bool(kp.is_difficult),
                        "is_exam_point": bool(kp.is_exam_point),
                    }
                    for kp in saved_kps
                ],
            }

        for idx, ch_node in enumerate(chapter_tree):
            created = await _create_node(None, idx + 1, ch_node)
            created_chapters.append(created)

        await session.commit()

        # 更新增量缓存标记
        course.content_hash = await _compute_course_content_hash(session, course_id)
        course.last_extracted_at = datetime.utcnow()
        await session.commit()

        # 统计知识点总数
        all_kp_count = 0
        def _count_kp(nodes):
            nonlocal all_kp_count
            for n in nodes:
                all_kp_count += len(n.get("knowledge_points", []))
                _count_kp(n.get("children", []))
        _count_kp(created_chapters)

        return ApiResponse(
            message=f"智能提取完成，已创建 {len(created_chapters)} 个章节，提取 {all_kp_count} 个知识点",
            data={
                "chapters": created_chapters,
                "total_kp": all_kp_count,
            },
        )


# ============================================================
# 知识点 CRUD（手动编辑/增删/标签切换）
# ============================================================
# 关系类型中文映射
_REL_TYPE_CN = {
    "依赖": "前置依赖",
    "支撑": "理论支撑",
    "组成": "组成部分",
    "对比": "对比关系",
    "应用": "应用关系",
    "prerequisite": "前置依赖",
}


def _kp_to_dict(kp: KnowledgePointORM) -> dict:
    return {
        "id": kp.id,
        "course_id": kp.course_id,
        "chapter_id": kp.chapter_id,
        "chapter": kp.chapter,
        "name": kp.name,
        "definition": kp.definition,
        "source_pages": kp.source_pages or "",
        "layer": kp.layer,
        "is_key_point": bool(kp.is_key_point),
        "is_difficult": bool(kp.is_difficult),
        "is_exam_point": bool(kp.is_exam_point),
        "prerequisites": kp.prerequisites_json or [],
        "relationships": kp.relationships_json or [],
        "sort_order": kp.sort_order,
    }


async def _build_textbook_context(
    session: AsyncSession,
    course_id: int,
    knowledge_points: list[dict],
) -> str:
    """检索多个知识点在教材中的原文上下文"""
    parts = []
    for kp in knowledge_points:
        kp_name = kp.get("name", "").strip()
        if not kp_name:
            continue
        try:
            ctx = await get_knowledge_point_context(session, course_id, kp_name)
        except Exception:
            continue
        if not ctx:
            continue
        context_text = ctx["context"]
        if len(context_text) > 1500:
            context_text = context_text[:1500] + "..."
        parts.append(f"知识点「{kp_name}」（教材第{ctx['page_number']}页附近）：\n{context_text}")
    return "\n\n---\n\n".join(parts)


@router.get("/courses/{course_id}/knowledge-points")
async def list_knowledge_points(course_id: int, chapter: Optional[str] = None, chapter_id: Optional[int] = None):
    """列出课程（或某章节）的知识点，支持按 chapter 名称或 chapter_id 筛选"""
    async for session in get_session():
        stmt = select(KnowledgePointORM).where(KnowledgePointORM.course_id == course_id)
        if chapter:
            stmt = stmt.where(KnowledgePointORM.chapter == chapter)
        if chapter_id is not None:
            stmt = stmt.where(KnowledgePointORM.chapter_id == chapter_id)
        stmt = stmt.order_by(KnowledgePointORM.sort_order)
        result = await session.execute(stmt)
        kps = result.scalars().all()
        return ApiResponse(data=[_kp_to_dict(k) for k in kps])


@router.post("/courses/{course_id}/knowledge-points")
async def create_knowledge_point(course_id: int, payload: dict):
    """手动新增知识点"""
    name = (payload.get("name") or "").strip()
    if not name:
        raise HTTPException(400, "知识点名称不能为空")
    chapter = (payload.get("chapter") or "").strip()

    async for session in get_session():
        course = await session.get(CourseORM, course_id)
        if not course:
            raise HTTPException(404, "课程不存在")
        # 计算 sort_order
        max_order = (await session.execute(
            select(func.max(KnowledgePointORM.sort_order))
            .where(KnowledgePointORM.course_id == course_id, KnowledgePointORM.chapter == chapter)
        )).scalar() or 0
        kp = KnowledgePointORM(
            course_id=course_id,
            chapter=chapter,
            chapter_id=payload.get("chapter_id"),
            name=name,
            definition=(payload.get("definition") or "").strip(),
            source_pages=(payload.get("source_pages") or "").strip(),
            layer=payload.get("layer") if payload.get("layer") in ("basic", "core", "extension") else "basic",
            is_key_point=1 if payload.get("is_key_point") else 0,
            is_difficult=1 if payload.get("is_difficult") else 0,
            is_exam_point=1 if payload.get("is_exam_point") else 0,
            prerequisites_json=payload.get("prerequisites") or [],
            relationships_json=payload.get("relationships") or [],
            sort_order=max_order + 1,
        )
        session.add(kp)
        await session.commit()
        await session.refresh(kp)
        return ApiResponse(data=_kp_to_dict(kp), message="知识点已新增")


@router.put("/knowledge-points/{kp_id}")
async def update_knowledge_point(kp_id: int, payload: dict):
    """更新知识点（名称/定义/层级/标签）"""
    async for session in get_session():
        kp = await session.get(KnowledgePointORM, kp_id)
        if not kp:
            raise HTTPException(404, "知识点不存在")
        if "name" in payload:
            name = (payload.get("name") or "").strip()
            if not name:
                raise HTTPException(400, "知识点名称不能为空")
            kp.name = name
        if "definition" in payload:
            kp.definition = (payload.get("definition") or "").strip()
        if "source_pages" in payload:
            kp.source_pages = (payload.get("source_pages") or "").strip()
        if "layer" in payload:
            kp.layer = payload.get("layer") if payload.get("layer") in ("basic", "core", "extension") else "basic"
        if "is_key_point" in payload:
            kp.is_key_point = 1 if payload.get("is_key_point") else 0
        if "is_difficult" in payload:
            kp.is_difficult = 1 if payload.get("is_difficult") else 0
        if "is_exam_point" in payload:
            kp.is_exam_point = 1 if payload.get("is_exam_point") else 0
        if "prerequisites" in payload:
            kp.prerequisites_json = payload.get("prerequisites") or []
        if "relationships" in payload:
            kp.relationships_json = payload.get("relationships") or []
        await session.commit()
        await session.refresh(kp)
        return ApiResponse(data=_kp_to_dict(kp), message="知识点已更新")


@router.delete("/knowledge-points/{kp_id}")
async def delete_knowledge_point(kp_id: int):
    """删除知识点"""
    async for session in get_session():
        kp = await session.get(KnowledgePointORM, kp_id)
        if not kp:
            raise HTTPException(404, "知识点不存在")
        await session.delete(kp)
        await session.commit()
        return ApiResponse(message="知识点已删除")


@router.get("/courses/{course_id}/knowledge-graph")
async def get_knowledge_graph(course_id: int):
    """获取课程知识图谱：所有知识点 + 前置关系（nodes + edges）"""
    async for session in get_session():
        course = await session.get(CourseORM, course_id)
        if not course:
            raise HTTPException(404, "课程不存在")
        stmt = select(KnowledgePointORM).where(
            KnowledgePointORM.course_id == course_id
        ).order_by(KnowledgePointORM.layer, KnowledgePointORM.sort_order)
        result = await session.execute(stmt)
        kps = result.scalars().all()

        name_map = {}  # 知识点名称 -> 知识点对象
        nodes = []
        for kp in kps:
            d = _kp_to_dict(kp)
            nodes.append(d)
            name_map[kp.name.lower()] = d

        edges = []
        for kp in kps:
            prereqs = kp.prerequisites_json or []
            for prereq_name in prereqs:
                target = name_map.get(prereq_name.strip().lower())
                if target:
                    edges.append({
                        "source": prereq_name.strip(),
                        "target": kp.name,
                        "rel_type": "prerequisite",
                        "label": "前置依赖",
                    })

            rels = kp.relationships_json or []
            for rel in rels:
                target_name = (rel.get("target") or "").strip()
                rel_type = (rel.get("rel_type") or "依赖").strip()
                if target_name and name_map.get(target_name.lower()):
                    edges.append({
                        "source": target_name,
                        "target": kp.name,
                        "rel_type": rel_type,
                        "label": _REL_TYPE_CN.get(rel_type, rel_type),
                    })

        return ApiResponse(data={
            "course_id": course_id,
            "course_name": course.name,
            "nodes": nodes,
            "edges": edges,
            "node_count": len(nodes),
            "edge_count": len(edges),
        })


# ============================================================
# 思维导图生成
# ============================================================
@router.post("/courses/{course_id}/mindmap")
async def generate_mindmap(course_id: int, payload: dict):
    """根据知识点生成思维导图数据（markmap 格式）"""
    async for session in get_session():
        course = await session.get(CourseORM, course_id)
        if not course:
            raise HTTPException(404, "课程不存在")

        # 获取知识点：优先使用请求传入的，否则从数据库取
        points_input = payload.get("knowledge_points")
        if points_input and isinstance(points_input, list):
            kps = points_input
        else:
            chapter = payload.get("chapter", "")
            stmt = select(KnowledgePointORM).where(KnowledgePointORM.course_id == course_id)
            if chapter:
                stmt = stmt.where(KnowledgePointORM.chapter == chapter)
            stmt = stmt.order_by(KnowledgePointORM.sort_order)
            result = await session.execute(stmt)
            kps = result.scalars().all()
            kps = [
                {
                    "name": k.name,
                    "definition": k.definition,
                    "layer": k.layer,
                    "is_key_point": bool(k.is_key_point),
                    "is_difficult": bool(k.is_difficult),
                    "is_exam_point": bool(k.is_exam_point),
                }
                for k in kps
            ]

        if not kps:
            raise HTTPException(400, "暂无知识点，请先提取知识点")

        # 按层级分组
        layers = {"basic": [], "core": [], "extension": []}
        for p in kps:
            layer = p.get("layer", "basic")
            if layer not in layers:
                layer = "basic"
            layers[layer].append(p)

        layer_labels = {"basic": "基础层", "core": "核心层", "extension": "拓展层"}
        layer_icons = {"basic": "📘", "core": "📗", "extension": "📕"}

        # 生成 markmap 兼容的 markdown
        lines = [f"# {course.name}"]
        chapter_name = payload.get("chapter", kps[0].get("chapter", "")) if kps else ""
        if chapter_name:
            lines[0] = f"# {course.name} — {chapter_name}"

        for layer_key in ("basic", "core", "extension"):
            items = layers[layer_key]
            if not items:
                continue
            lines.append(f"## {layer_icons[layer_key]} {layer_labels[layer_key]} ({len(items)}个)")
            for p in items:
                name = p.get("name", "")
                definition = p.get("definition", "")
                tags = []
                if p.get("is_key_point"):
                    tags.append("重点")
                if p.get("is_difficult"):
                    tags.append("难点")
                if p.get("is_exam_point"):
                    tags.append("考点")
                tag_str = f" [{', '.join(tags)}]" if tags else ""
                desc = f"：{definition[:80]}" if definition else ""
                lines.append(f"### {name}{tag_str}{desc}")

        markdown = "\n".join(lines)

        return ApiResponse(
            message="思维导图生成成功",
            data={
                "markdown": markdown,
                "count": len(kps),
            },
        )


async def _apply_template_defaults(plan_dict: dict, template_structure: dict | None, course_name: str, chapter: str) -> dict:
    """将模板 defaults 回填到 plan_dict，缺失字段自动补齐（保证全表格可渲染）"""
    if not template_structure:
        return plan_dict
    defaults = template_structure.get("defaults") or {}
    merged = {**defaults, **plan_dict}
    # 固定优先值
    merged["course_name"] = course_name or merged.get("course_name", "")
    merged["chapter"] = chapter or merged.get("chapter", "")
    # stages 长度校验
    if not merged.get("stages") or not isinstance(merged["stages"], list):
        merged["stages"] = defaults.get("stages", [])
    # key_points / difficult_points / homework 保证是 list
    for k in ("key_points", "difficult_points", "homework"):
        if not isinstance(merged.get(k), list):
            merged[k] = defaults.get(k, [])
    # 知识点出处表（knowledge_sources）默认保证为空列表
    if "knowledge_sources" not in merged or not isinstance(merged["knowledge_sources"], list):
        merged["knowledge_sources"] = defaults.get("knowledge_sources", [])
    return merged


# ============================================================
# 教案生成
# ============================================================
@router.post("/courses/{course_id}/generate-lesson")
async def generate_lesson_api(
    course_id: int,
    chapter: str = Form(...),
    knowledge_points: str = Form(...),  # JSON字符串
    params: Optional[str] = Form(None),  # JSON字符串，可选
    chapter_id: Optional[int] = Form(None),  # 关联章节树节点
    template_id: Optional[int] = Form(None),  # 教案模板ID
    mode: Optional[str] = Form(None),  # 生成模式: "addie" 启用多智能体审议; 默认/其他 fast模式
):
    async for session in get_session():
        course = await session.get(CourseORM, course_id)
        if not course:
            raise HTTPException(404, "课程不存在")

        # 加载模板结构（template_id or 默认）
        template_structure: dict = {}
        if template_id:
            tmpl = await session.get(LessonTemplateORM, template_id)
            if tmpl:
                template_structure = tmpl.structure_json or {}
        if not template_structure:
            r = await session.execute(select(LessonTemplateORM).where(LessonTemplateORM.is_default == True).limit(1))
            default_tmpl = r.scalars().first()
            if default_tmpl:
                template_structure = default_tmpl.structure_json or {}

        # 解析知识点
        try:
            kp_list = json.loads(knowledge_points)
            if not isinstance(kp_list, list):
                raise ValueError
        except Exception:
            raise HTTPException(400, "knowledge_points 必须是JSON数组")

        # 解析参数
        try:
            params_dict = json.loads(params) if params else {}
            lesson_params = LessonParams(**params_dict)
        except Exception as e:
            raise HTTPException(400, f"参数解析失败: {e}")

        # 检索教材原文作为上下文
        textbook_context_parts = []
        for kp in kp_list:
            kp_name = kp.get("name", "").strip()
            if kp_name:
                ctx = await get_knowledge_point_context(session, course_id, kp_name)
                if ctx:
                    context_text = ctx["context"]
                    if len(context_text) > 1500:
                        context_text = context_text[:1500] + "..."
                    textbook_context_parts.append(
                        f"知识点「{kp_name}」（教材第{ctx['page_number']}页附近）：\n{context_text}"
                    )
        textbook_context = "\n\n---\n\n".join(textbook_context_parts)

        # 生成教案：捕获异常 → 结构化返回 + 模板骨架 fallback
        plan_dict: dict = {}
        addie_meta: dict = {}
        error_code = None
        error_message = None
        try:
            if (mode or "").strip().lower() == "addie":
                # ADDIE 多智能体审议流程：4 阶段 (Analyze → Develop → Evaluate → Refine)
                raw_lesson_dict, addie_meta = await generate_lesson_with_addie(
                    course.name, chapter, kp_list, lesson_params,
                    textbook_context=textbook_context,
                    subject=course.subject or None,
                )
                # 复用 LessonPlan 模型做结构校验与字段标准化
                from ..models.schemas import LessonPlan as _LessonPlan
                plan = _LessonPlan(**raw_lesson_dict)
                plan_dict = plan.model_dump()
            else:
                plan = await generate_lesson(
                    course.name, chapter, kp_list, lesson_params,
                    textbook_context=textbook_context,
                    subject=course.subject or None,
                )
                plan_dict = plan.model_dump()
        except LLMError as e:
            error_code = "LLM_ERROR"
            error_message = f"AI服务调用失败: {e}"
        except Exception as e:
            error_code = "GENERATE_UNKNOWN_ERROR"
            error_message = f"教案生成异常: {type(e).__name__}: {e}"

        # 生成失败 → 用模板默认结构兜底（保证用户拿到可编辑的完整骨架）
        if error_code:
            plan_dict = await _apply_template_defaults(plan_dict, template_structure, course.name, chapter)
            # 构造知识出处
            plan_dict["knowledge_sources"] = [
                {"name": kp.get("name",""), "source": kp.get("source_pages","") or "教材原文", "description": kp.get("definition","")[:80]}
                for kp in kp_list if kp.get("name")
            ]
            lesson = LessonORM(
                course_id=course_id,
                chapter_id=chapter_id,
                chapter=chapter,
                title=f"{course.name} - {chapter}（AI失败·模板骨架）",
                plan_json=plan_dict,
                params_json=lesson_params.model_dump(),
            )
            session.add(lesson)
            await session.commit()
            await session.refresh(lesson)
            return ApiResponse(
                success=False,
                message=error_message or "AI生成失败，已使用模板生成骨架",
                data={
                    "error_code": error_code,
                    "fallbacks": FALLBACK_ACTIONS,
                    "fallback_used": "template_fallback",
                    "id": lesson.id,
                    "course_id": lesson.course_id,
                    "chapter": lesson.chapter,
                    "title": lesson.title,
                    "plan": plan_dict,
                    "params": lesson_params.model_dump(),
                    "created_at": lesson.created_at.isoformat() if lesson.created_at else None,
                    "updated_at": lesson.updated_at.isoformat() if lesson.updated_at else None,
                },
            )

        # 正常：应用模板补齐
        plan_dict = await _apply_template_defaults(plan_dict, template_structure, course.name, chapter)
        # 补齐知识点出处（根据传入的KP + source_pages）
        if not plan_dict.get("knowledge_sources"):
            plan_dict["knowledge_sources"] = [
                {"name": kp.get("name",""), "source": kp.get("source_pages","") or "教材原文", "description": kp.get("definition","")[:80]}
                for kp in kp_list if kp.get("name")
            ]

        # 内容后置校验（借鉴 AgentCourseAssistant 的 ContentValidator）
        # 仅产出 warnings/errors 提示，不阻塞入库
        validation_result = validate_lesson(plan_dict)

        # 入库
        lesson = LessonORM(
            course_id=course_id,
            chapter_id=chapter_id,
            chapter=chapter,
            title=f"{course.name} - {chapter}",
            plan_json=plan_dict,
            params_json=lesson_params.model_dump(),
        )
        session.add(lesson)
        await session.commit()
        await session.refresh(lesson)

        return ApiResponse(
            message="教案生成成功" + ("（ADDIE 多智能体审议）" if addie_meta else ""),
            data={
                "id": lesson.id,
                "course_id": lesson.course_id,
                "chapter": lesson.chapter,
                "title": lesson.title,
                "plan": plan.model_dump(),
                "params": lesson_params.model_dump(),
                "validation": validation_result.to_dict(),
                "addie_meta": addie_meta or None,
                "created_at": lesson.created_at.isoformat() if lesson.created_at else None,
                "updated_at": lesson.updated_at.isoformat() if lesson.updated_at else None,
            },
        )


# ============================================================
# 教案管理
# ============================================================
@router.post("/courses/{course_id}/lessons")
async def create_lesson_manual(course_id: int, payload: dict):
    """手动创建空白教案（供后端先建记录再调用 generate / fallback-template）"""
    async for session in get_session():
        course = await session.get(CourseORM, course_id)
        if not course:
            raise HTTPException(404, "课程不存在")
        section_id = payload.get("section_id") or payload.get("chapter_id")
        chapter_title = (payload.get("chapter") or payload.get("section") or payload.get("title") or "未命名章节").strip()[:200]
        title = (payload.get("title") or chapter_title)[:200]
        duration = int(payload.get("duration_minutes") or 45)
        if section_id:
            chapter_row = await session.get(ChapterORM, int(section_id))
            if chapter_row:
                chapter_title = chapter_row.name
        lesson = LessonORM(
            course_id=course_id,
            chapter_id=int(section_id) if section_id else None,
            chapter=chapter_title,
            title=title,
            plan_json={},
            params_json={"duration_minutes": duration},
            source_material_ids=[],
        )
        session.add(lesson)
        await session.commit()
        await session.refresh(lesson)
        return ApiResponse(message="教案记录已创建", data={
            "id": lesson.id, "title": lesson.title, "chapter": lesson.chapter,
            "chapter_id": lesson.chapter_id,
        })


@router.get("/courses/{course_id}/lessons")
async def list_lessons(course_id: int):
    async for session in get_session():
        result = await session.execute(
            select(LessonORM)
            .where(LessonORM.course_id == course_id)
            .order_by(LessonORM.updated_at.desc())
        )
        lessons = result.scalars().all()
        return ApiResponse(
            data=[
                {
                    "id": l.id,
                    "course_id": l.course_id,
                    "chapter": l.chapter,
                    "title": l.title,
                    "created_at": l.created_at.isoformat() if l.created_at else None,
                    "updated_at": l.updated_at.isoformat() if l.updated_at else None,
                }
                for l in lessons
            ]
        )


@router.get("/lessons/{lesson_id}")
async def get_lesson(lesson_id: int):
    async for session in get_session():
        lesson = await session.get(LessonORM, lesson_id)
        if not lesson:
            raise HTTPException(404, "教案不存在")
        return ApiResponse(
            data={
                "id": lesson.id,
                "course_id": lesson.course_id,
                "chapter": lesson.chapter,
                "title": lesson.title,
                "plan": lesson.plan_json,
                "params": lesson.params_json,
                "created_at": lesson.created_at.isoformat() if lesson.created_at else None,
                "updated_at": lesson.updated_at.isoformat() if lesson.updated_at else None,
            }
        )


@router.put("/lessons/{lesson_id}")
async def update_lesson(lesson_id: int, plan: dict):
    """直接更新教案（用于编辑器修改后保存）"""
    async for session in get_session():
        lesson = await session.get(LessonORM, lesson_id)
        if not lesson:
            raise HTTPException(404, "教案不存在")
        lesson.plan_json = plan
        lesson.updated_at = datetime.utcnow()
        await session.commit()
        return ApiResponse(message="教案已保存")


@router.post("/lessons/{lesson_id}/evaluate")
async def evaluate_lesson_api(lesson_id: int):
    """对已生成教案进行独立质量评估

    借鉴 instructional_agents/evaluate.py 的多指标打分+双视角评审机制：
    - 6 维度细粒度打分（结构/目标/重难点/学情/教学法/可执行性）
    - 教务专家视角 + 学生代表视角 双重评审
    返回 JSON 评估报告。
    """
    async for session in get_session():
        lesson = await session.get(LessonORM, lesson_id)
        if not lesson:
            raise HTTPException(404, "教案不存在")
        plan_dict = lesson.plan_json or {}
        if not plan_dict.get("stages"):
            raise HTTPException(400, "教案内容为空，无法评估")
        course = await session.get(CourseORM, lesson.course_id)
        course_name = course.name if course else "未命名课程"

        # 拉取本节关联知识点（用于交叉校验目标达成度）
        kp_result = await session.execute(
            select(KnowledgePointORM)
            .where(KnowledgePointORM.course_id == lesson.course_id)
            .where(KnowledgePointORM.chapter == lesson.chapter)
        )
        kp_rows = kp_result.scalars().all()
        knowledge_points = [
            {"name": k.name, "source_pages": k.source_pages or "", "definition": k.definition or ""}
            for k in kp_rows
        ] or []

        try:
            subject = (course.subject or None) if course else None
            result = await evaluate_lesson_independent(
                course_name=course_name,
                chapter=lesson.chapter,
                lesson_dict=plan_dict,
                knowledge_points=knowledge_points,
                subject=subject,
            )
        except LLMError as e:
            return ApiResponse(
                success=False,
                message=f"AI服务调用失败: {e}",
                data={"error_code": "LLM_ERROR", "fallbacks": FALLBACK_ACTIONS},
            )
        except Exception as e:
            return ApiResponse(
                success=False,
                message=f"评估异常: {type(e).__name__}: {e}",
                data={"error_code": "EVALUATE_UNKNOWN_ERROR", "fallbacks": FALLBACK_ACTIONS},
            )

        return ApiResponse(
            message=f"教案评估完成（总分 {result.get('overall_score', 0):.1f}/5.0）",
            data=result,
        )


@router.delete("/lessons/{lesson_id}")
async def delete_lesson(lesson_id: int):
    async for session in get_session():
        lesson = await session.get(LessonORM, lesson_id)
        if not lesson:
            raise HTTPException(404, "教案不存在")
        await session.delete(lesson)
        await session.commit()
        return ApiResponse(message="教案已删除")


# ============================================================
# 本地上传教案 / PPT
# ============================================================
@router.post("/courses/{course_id}/upload-lesson")
async def upload_lesson_file(course_id: int, file: UploadFile = File(...)):
    """上传本地教案文件（DOCX/MD/TXT），自动解析并创建教案记录"""
    ext = Path(file.filename).suffix.lower() if file.filename else ""
    if ext not in (".docx", ".md", ".txt"):
        raise HTTPException(400, "仅支持 .docx/.md/.txt 格式教案文件")

    async for session in get_session():
        course = await session.get(CourseORM, course_id)
        if not course:
            raise HTTPException(404, "课程不存在")

        # 读取文件内容
        content_bytes = await file.read()
        content_text = content_bytes.decode("utf-8", errors="ignore")

        # 保存文件
        stored_path = await save_upload(content_bytes, file.filename or "lesson_upload.txt", course_id)

        # 构建基础教案结构
        chapter_name = Path(file.filename).stem[:30]
        plan = {
            "course_name": course.name,
            "chapter": chapter_name,
            "total_minutes": 90,
            "knowledge_goal": "",
            "ability_goal": "",
            "value_goal": "",
            "key_points": [],
            "difficult_points": [],
            "difficult_strategy": "",
            "stages": [],
            "homework": [],
            "board_design": "",
            "reflection": "（课后填写）",
            "source_text": content_text[:50000],
        }

        lesson = LessonORM(
            course_id=course_id,
            chapter=chapter_name,
            title=chapter_name,
            plan_json=plan,
            params_json={},
        )
        session.add(lesson)
        await session.flush()
        lesson_id = lesson.id
        await session.commit()

        return ApiResponse(
            data={
                "id": lesson_id,
                "course_id": course_id,
                "chapter": chapter_name,
                "title": chapter_name,
                "created_at": lesson.created_at.isoformat() if lesson.created_at else None,
                "uploaded": True,
            }
        )


@router.post("/courses/{course_id}/upload-ppt")
async def upload_ppt_file(course_id: int, file: UploadFile = File(...)):
    """上传本地PPT文件（PPTX），存储到课程"""
    ext = Path(file.filename).suffix.lower() if file.filename else ""
    if ext != ".pptx":
        raise HTTPException(400, "仅支持 .pptx 格式PPT文件")

    async for session in get_session():
        course = await session.get(CourseORM, course_id)
        if not course:
            raise HTTPException(404, "课程不存在")

        content_bytes = await file.read()
        stored_path = await save_upload(content_bytes, file.filename or "upload.pptx", course_id)

        title = Path(file.filename).stem[:30]
        ppt = PptRecordORM(
            course_id=course_id,
            chapter=title,
            title=title,
            slide_count=0,
            slide_data={},
            source="upload",
            stored_path=str(stored_path),
        )
        session.add(ppt)
        await session.flush()
        ppt_id = ppt.id
        await session.commit()

        return ApiResponse(
            data={
                "id": ppt_id,
                "course_id": course_id,
                "title": title,
                "source": "upload",
                "created_at": ppt.created_at.isoformat() if ppt.created_at else None,
            }
        )


@router.get("/ppt/{ppt_id}/download")
async def download_ppt_file(ppt_id: int):
    """下载上传的PPT文件"""
    async for session in get_session():
        ppt = await session.get(PptRecordORM, ppt_id)
        if not ppt:
            raise HTTPException(404, "PPT记录不存在")
        if not ppt.stored_path:
            raise HTTPException(404, "该PPT没有关联文件")

        file_path = Path(ppt.stored_path)
        if not file_path.exists():
            raise HTTPException(404, "文件已被删除")

        content = file_path.read_bytes()
        filename_encoded = urllib.parse.quote(f"{ppt.title}.pptx")
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
            headers={
                "Content-Disposition": f"attachment; filename=\"{ppt.title}.pptx\"; filename*=UTF-8''{filename_encoded}"
            },
        )


# ============================================================
# 对话修改
# ============================================================
@router.get("/courses/{course_id}/chat-messages")
async def list_course_chat_messages(course_id: int, limit: int = 200):
    """获取指定课程的全部聊天记录（按课程为单位保留对话）
    返回按时间正序的 [{id, role, content, created_at, metadata}] 列表"""
    async for session in get_session():
        stmt = (
            select(ChatMessageORM)
            .where(ChatMessageORM.course_id == course_id)
            .order_by(ChatMessageORM.created_at.asc(), ChatMessageORM.id.asc())
            .limit(min(max(limit, 1), 500))
        )
        result = await session.execute(stmt)
        msgs = result.scalars().all()
        return ApiResponse(
            data=[
                {
                    "id": m.id,
                    "role": m.role,
                    "content": m.content,
                    "created_at": m.created_at.isoformat() if m.created_at else None,
                    "metadata": m.metadata_json or {},
                }
                for m in msgs
            ]
        )


@router.post("/lessons/{lesson_id}/chat")
async def chat_modify_lesson(lesson_id: int, payload: ChatRequest):
    """通过自然语言修改教案"""
    async for session in get_session():
        lesson = await session.get(LessonORM, lesson_id)
        if not lesson:
            raise HTTPException(404, "教案不存在")

        # 记录用户消息
        session.add(
            ChatMessageORM(
                course_id=lesson.course_id,
                role="user",
                content=payload.message,
            )
        )

        try:
            plan = LessonPlan(**lesson.plan_json)
        except Exception as e:
            raise HTTPException(500, f"教案数据损坏: {e}")

        try:
            result = await modify_lesson(
                course_name=plan.course_name,
                chapter=plan.chapter,
                current_plan=plan,
                user_message=payload.message,
            )
        except LLMError as e:
            raise HTTPException(502, str(e))

        response_text = ""
        new_plan_dict = None

        if result["type"] == "modified":
            new_plan: LessonPlan = result["plan"]
            lesson.plan_json = new_plan.model_dump()
            lesson.updated_at = datetime.utcnow()
            new_plan_dict = new_plan.model_dump()
            response_text = "已根据您的指令修改教案。"
        elif result["type"] == "clarify":
            response_text = result["question"]
        else:  # answer
            response_text = result["answer"]

        # 记录助手消息
        session.add(
            ChatMessageORM(
                course_id=lesson.course_id,
                role="assistant",
                content=response_text,
                metadata_json={"type": result["type"]},
            )
        )
        await session.commit()

        return ApiResponse(
            data={
                "type": result["type"],
                "response": response_text,
                "plan": new_plan_dict,
            }
        )


# ============================================================
# 导出
# ============================================================
@router.get("/lessons/{lesson_id}/export/{fmt}")
async def export_lesson(lesson_id: int, fmt: str):
    """导出教案：fmt=markdown|docx|pptx"""
    if fmt not in ("markdown", "docx", "pptx"):
        raise HTTPException(400, "格式仅支持 markdown、docx 或 pptx")

    async for session in get_session():
        lesson = await session.get(LessonORM, lesson_id)
        if not lesson:
            raise HTTPException(404, "教案不存在")
        try:
            plan = LessonPlan(**lesson.plan_json)
        except Exception as e:
            raise HTTPException(500, f"教案数据损坏: {e}")

        # 安全的文件名：课程名+章节，去除特殊字符
        safe_title = f"{plan.course_name}_{plan.chapter}".replace(" ", "_")
        safe_title = "".join(c for c in safe_title if c.isalnum() or c in "_-")
        if not safe_title:
            safe_title = f"lesson_{lesson_id}"

        # ASCII fallback 文件名（header 不支持非 ASCII）
        ascii_fallback = f"lesson_{lesson_id}"

        if fmt == "markdown":
            content = export_markdown(plan)
            filename_encoded = urllib.parse.quote(f"{safe_title}.md")
            return Response(
                content=content.encode("utf-8"),
                media_type="text/markdown; charset=utf-8",
                headers={
                    "Content-Disposition": f"attachment; filename=\"{ascii_fallback}.md\"; filename*=UTF-8''{filename_encoded}"
                },
            )
        elif fmt == "docx":
            try:
                content = export_docx(plan)
            except Exception as e:
                raise HTTPException(500, f"DOCX生成失败: {e}")
            filename_encoded = urllib.parse.quote(f"{safe_title}.docx")
            return Response(
                content=content,
                media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                headers={
                    "Content-Disposition": f"attachment; filename=\"{ascii_fallback}.docx\"; filename*=UTF-8''{filename_encoded}"
                },
            )
        else:  # pptx — 使用默认参数生成教学PPT
            # 获取课程以读取学科(用于注入学科领域规则)
            course = await session.get(CourseORM, lesson.course_id)
            course_subject = course.subject if course else ""
            # 获取知识点
            stmt = (
                select(KnowledgePointORM)
                .where(KnowledgePointORM.course_id == lesson.course_id)
                .order_by(KnowledgePointORM.sort_order)
            )
            result = await session.execute(stmt)
            kps = result.scalars().all()
            knowledge_points = [
                {
                    "name": k.name,
                    "definition": k.definition,
                    "source_pages": k.source_pages or "",
                    "layer": k.layer,
                    "is_key_point": bool(k.is_key_point),
                    "is_difficult": bool(k.is_difficult),
                    "is_exam_point": bool(k.is_exam_point),
                }
                for k in kps
            ]

            # 检索教材原文上下文
            textbook_context = await _build_textbook_context(session, lesson.course_id, knowledge_points)

            try:
                slide_data = await generate_ppt_content(
                    plan=plan,
                    knowledge_points=knowledge_points,
                    style="cyan_ink",
                    content_density="moderate",
                    image_style="icons",
                    textbook_context=textbook_context,
                    subject=course_subject or None,
                )
                content = export_teaching_pptx(slide_data, style="cyan_ink")
            except Exception as e:
                raise HTTPException(500, f"PPT生成失败: {e}")
            filename_encoded = urllib.parse.quote(f"{safe_title}.pptx")
            return Response(
                content=content,
                media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                headers={
                    "Content-Disposition": f"attachment; filename=\"{ascii_fallback}.pptx\"; filename*=UTF-8''{filename_encoded}"
                },
            )


@router.post("/lessons/{lesson_id}/export-ppt")
async def export_lesson_ppt(lesson_id: int, payload: dict):
    """生成教学PPT（带参数配置）

    请求体：
    {
        "style": "academic|cyan_ink|cute_cartoon|formal|minimal",
        "content_density": "detailed|moderate|concise",
        "image_style": "none|icons|rich",
        "style_custom": "用户自定义风格描述（可选）"
    }
    """
    style = payload.get("style", "cyan_ink")
    content_density = payload.get("content_density", "moderate")
    image_style = payload.get("image_style", "icons")
    style_custom = payload.get("style_custom", "")

    async for session in get_session():
        lesson = await session.get(LessonORM, lesson_id)
        if not lesson:
            raise HTTPException(404, "教案不存在")
        try:
            plan = LessonPlan(**lesson.plan_json)
        except Exception as e:
            raise HTTPException(500, f"教案数据损坏: {e}")

        # 获取课程以读取学科(用于注入学科领域规则)
        course = await session.get(CourseORM, lesson.course_id)
        course_subject = course.subject if course else ""

        # 获取知识点
        stmt = (
            select(KnowledgePointORM)
            .where(KnowledgePointORM.course_id == lesson.course_id)
            .order_by(KnowledgePointORM.sort_order)
        )
        result = await session.execute(stmt)
        kps = result.scalars().all()
        knowledge_points = [
            {
                "name": k.name,
                "definition": k.definition,
                "source_pages": k.source_pages or "",
                "layer": k.layer,
                "is_key_point": bool(k.is_key_point),
                "is_difficult": bool(k.is_difficult),
                "is_exam_point": bool(k.is_exam_point),
            }
            for k in kps
        ]

        # 检索教材原文上下文
        textbook_context = await _build_textbook_context(session, lesson.course_id, knowledge_points)

        try:
            slide_data = await generate_ppt_content(
                plan=plan,
                knowledge_points=knowledge_points,
                style=style,
                content_density=content_density,
                image_style=image_style,
                style_custom=style_custom,
                textbook_context=textbook_context,
                subject=course_subject or None,
            )
            content = export_teaching_pptx(slide_data, style=style)
        except LLMError as e:
            # AI失败：先尝试 minimal 风格降参数生成
            try:
                slide_data = await generate_ppt_content(
                    plan=plan,
                    knowledge_points=knowledge_points,
                    style="minimal",
                    content_density="concise",
                    image_style="none",
                    style_custom="",
                    textbook_context=textbook_context,
                    subject=course_subject or None,
                )
                content = export_teaching_pptx(slide_data, style="minimal")
            except Exception as e2:
                return ApiResponse(
                    success=False,
                    message=f"PPT AI生成失败，且降级生成也失败: {e2}",
                    data={
                        "error_code": "PPT_LLM_AND_FALLBACK_FAILED",
                        "fallbacks": FALLBACK_ACTIONS,
                        "detail_primary": str(e),
                        "detail_fallback": str(e2),
                    }
                ).model_dump()
        except Exception as e:
            return ApiResponse(
                success=False,
                message=f"PPT生成失败: {e}",
                data={
                    "error_code": "PPT_GENERATE_ERROR",
                    "fallbacks": FALLBACK_ACTIONS,
                }
            ).model_dump()

        # 保存PPT记录
        slide_count = 0
        if slide_data and isinstance(slide_data, dict):
            slides = slide_data.get("slides", [])
            if isinstance(slides, list):
                slide_count = len(slides)

        # 内容后置校验（借鉴 AgentCourseAssistant 的 ContentValidator）
        # 校验结果通过 HTTP header 返回，不阻塞文件下载
        ppt_validation = validate_ppt(slide_data) if isinstance(slide_data, dict) else None

        ppt_record = PptRecordORM(
            course_id=lesson.course_id,
            lesson_id=lesson_id,
            chapter=plan.chapter,
            title=f"{plan.course_name} - {plan.chapter}",
            style=style,
            content_density=content_density,
            image_style=image_style,
            style_custom=style_custom,
            slide_count=slide_count,
            slide_data=slide_data,
        )
        session.add(ppt_record)
        await session.commit()
        await session.refresh(ppt_record)

        # 安全的文件名
        safe_title = f"{plan.course_name}_{plan.chapter}".replace(" ", "_")
        safe_title = "".join(c for c in safe_title if c.isalnum() or c in "_-")
        if not safe_title:
            safe_title = f"lesson_{lesson_id}"
        ascii_fallback = f"lesson_{lesson_id}"
        filename_encoded = urllib.parse.quote(f"{safe_title}.pptx")
        # 将校验结果通过 header 返回（前端可读取 X-Ppt-Validation）
        ppt_validation_header = ""
        if ppt_validation is not None:
            try:
                ppt_validation_header = urllib.parse.quote(
                    json.dumps(ppt_validation.to_dict(), ensure_ascii=False)
                )
            except Exception:
                pass
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
            headers={
                "Content-Disposition": f"attachment; filename=\"{ascii_fallback}.pptx\"; filename*=UTF-8''{filename_encoded}",
                "X-Ppt-Record-Id": str(ppt_record.id),
                "X-Ppt-Validation": ppt_validation_header or "{}",
            },
        )


# ============================================================
# PPT记录管理
# ============================================================
@router.get("/courses/{course_id}/ppt-records")
async def list_ppt_records(course_id: int):
    """列出课程的PPT生成记录"""
    async for session in get_session():
        result = await session.execute(
            select(PptRecordORM)
            .where(PptRecordORM.course_id == course_id)
            .order_by(PptRecordORM.created_at.desc())
        )
        records = result.scalars().all()
        return ApiResponse(
            data=[
                {
                    "id": r.id,
                    "course_id": r.course_id,
                    "lesson_id": r.lesson_id,
                    "chapter": r.chapter,
                    "title": r.title,
                    "style": r.style,
                    "content_density": r.content_density,
                    "image_style": r.image_style,
                    "slide_count": r.slide_count,
                    "source": r.source,
                    "has_file": bool(r.stored_path),
                    "created_at": r.created_at.isoformat() if r.created_at else None,
                }
                for r in records
            ]
        )


@router.get("/ppt-records/{record_id}")
async def get_ppt_record(record_id: int):
    """获取PPT记录详情"""
    async for session in get_session():
        r = await session.get(PptRecordORM, record_id)
        if not r:
            raise HTTPException(404, "PPT记录不存在")
        return ApiResponse(
            data={
                "id": r.id,
                "course_id": r.course_id,
                "lesson_id": r.lesson_id,
                "chapter": r.chapter,
                "title": r.title,
                "style": r.style,
                "content_density": r.content_density,
                "image_style": r.image_style,
                "style_custom": r.style_custom or "",
                "slide_count": r.slide_count,
                "slide_data": r.slide_data,
                "created_at": r.created_at.isoformat() if r.created_at else None,
            }
        )


@router.delete("/ppt-records/{record_id}")
async def delete_ppt_record(record_id: int):
    """删除PPT记录"""
    async for session in get_session():
        r = await session.get(PptRecordORM, record_id)
        if not r:
            raise HTTPException(404, "PPT记录不存在")
        await session.delete(r)
        await session.commit()
        return ApiResponse(message="PPT记录已删除")


@router.post("/ppt-records/{record_id}/download")
async def download_ppt(record_id: int):
    """重新下载PPT（根据已有记录生成）"""
    async for session in get_session():
        r = await session.get(PptRecordORM, record_id)
        if not r:
            raise HTTPException(404, "PPT记录不存在")

        slide_data = r.slide_data
        if not slide_data:
            raise HTTPException(400, "PPT数据为空，请重新生成")

        try:
            content = export_teaching_pptx(slide_data, style=r.style)
        except Exception as e:
            raise HTTPException(500, f"PPT生成失败: {e}")

        safe_title = r.title or f"ppt_{record_id}"
        safe_title = "".join(c for c in safe_title if c.isalnum() or c in "_-")
        if not safe_title:
            safe_title = f"ppt_{record_id}"
        ascii_fallback = f"ppt_{record_id}"
        filename_encoded = urllib.parse.quote(f"{safe_title}.pptx")

        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
            headers={
                "Content-Disposition": f"attachment; filename=\"{ascii_fallback}.pptx\"; filename*=UTF-8''{filename_encoded}"
            },
        )


# ============================================================
# 健康检查 / API连通性
# ============================================================
@router.get("/health")
async def health():
    return ApiResponse(message="ok", data={"status": "healthy"})


@router.get("/llm-test")
async def llm_test():
    """测试当前LLM连通性"""
    cfg = get_llm_config()
    if not cfg.is_configured():
        raise HTTPException(
            502,
            "LLM未配置，请点击右上角「设置」按钮配置 API Key、Base URL 和模型",
        )
    try:
        llm = get_llm()
        text = await llm.chat(
            system_prompt="你是测试助手。",
            user_prompt="请回复：连通正常",
            temperature=0.0,
            max_tokens=30,
        )
        current = cfg.get_masked()
        return ApiResponse(
            message="LLM连通正常",
            data={
                "response": text,
                "provider": current["provider"],
                "model": current["model"],
                "base_url": current["base_url"],
            },
        )
    except LLMError as e:
        raise HTTPException(502, f"LLM不可用: {e}")


# ============================================================
# LLM 设置管理
# ============================================================
@router.get("/settings/llm")
async def get_llm_settings():
    """获取当前 LLM 设置（API Key 掩码）"""
    cfg = get_llm_config()
    return ApiResponse(
        data={
            "current": cfg.get_masked(),
            "providers": PRESET_PROVIDERS,
            "is_configured": cfg.is_configured(),
        }
    )


@router.put("/settings/llm")
async def update_llm_settings(payload: LLMSettingsUpdate):
    """保存 LLM 设置"""
    cfg = get_llm_config()
    # 过滤 None 值
    update_data = {k: v for k, v in payload.model_dump().items() if v is not None}
    if not update_data:
        return ApiResponse(message="无更新内容", data=cfg.get_masked())

    try:
        cfg.update(update_data)
        reset_llm()
    except Exception as e:
        raise HTTPException(500, f"保存失败: {e}")

    return ApiResponse(
        message="设置已保存",
        data=cfg.get_masked(),
    )


@router.post("/settings/llm/test")
async def test_llm_settings(payload: LLMSettingsUpdate):
    """测试给定的 LLM 配置（不持久化，直接测试）"""
    cfg = get_llm_config()
    current = cfg.get_all()
    # 合并：用户提交的覆盖当前（处理掩码）
    test_cfg = {
        "provider": payload.provider or current["provider"],
        "api_key": current["api_key"],
        "base_url": payload.base_url or current["base_url"],
        "model": payload.model or current["model"],
        "temperature": payload.temperature if payload.temperature is not None else current["temperature"],
        "max_tokens": payload.max_tokens if payload.max_tokens is not None else current["max_tokens"],
    }
    # 若用户提交了非掩码 api_key，使用新值
    if payload.api_key and "****" not in payload.api_key:
        test_cfg["api_key"] = payload.api_key

    if not test_cfg["api_key"]:
        raise HTTPException(400, "请填写 API Key")
    if not test_cfg["base_url"]:
        raise HTTPException(400, "请填写 Base URL")
    if not test_cfg["model"]:
        raise HTTPException(400, "请填写模型名称")

    try:
        from ..core.llm import LLMClient

        client = LLMClient(test_cfg)
        text = await client.chat(
            system_prompt="你是测试助手。",
            user_prompt="请回复：连通正常",
            temperature=0.0,
            max_tokens=30,
        )
        return ApiResponse(
            message="连接测试成功",
            data={
                "response": text,
                "provider": test_cfg["provider"],
                "model": test_cfg["model"],
                "base_url": test_cfg["base_url"],
            },
        )
    except LLMError as e:
        raise HTTPException(502, f"连接测试失败: {e}")
    except Exception as e:
        raise HTTPException(502, f"连接测试失败: {type(e).__name__}: {e}")


# 引入settings用于llm-test
from ..config import settings  # noqa: E402


# ============================================================
# 教案模板库 CRUD
# ============================================================
@router.get("/lesson-templates")
async def list_lesson_templates(course_id: Optional[int] = None):
    """列出教案模板：全局模板(course_id is null) + 指定课程私有模板"""
    async for session in get_session():
        stmt = select(LessonTemplateORM).order_by(
            LessonTemplateORM.is_default.desc(),
            LessonTemplateORM.created_at.desc(),
        )
        if course_id:
            stmt = stmt.where(
                (LessonTemplateORM.course_id.is_(None)) | (LessonTemplateORM.course_id == course_id)
            )
        else:
            stmt = stmt.where(LessonTemplateORM.course_id.is_(None))
        result = await session.execute(stmt)
        tpls = result.scalars().all()
        return ApiResponse(
            data=[
                {
                    "id": t.id,
                    "course_id": t.course_id,
                    "name": t.name,
                    "description": t.description,
                    "structure_json": t.structure_json,
                    "is_default": t.is_default,
                    "created_at": t.created_at.isoformat() if t.created_at else None,
                    "updated_at": t.updated_at.isoformat() if t.updated_at else None,
                }
                for t in tpls
            ]
        )


@router.post("/lesson-templates")
async def create_lesson_template(payload: LessonTemplateCreate):
    """新建教案模板"""
    async for session in get_session():
        # 若设为默认，先取消其他默认
        if payload.is_default:
            r = await session.execute(select(LessonTemplateORM).where(LessonTemplateORM.is_default == True))
            for old in r.scalars().all():
                old.is_default = False
        tpl = LessonTemplateORM(
            course_id=payload.course_id,
            name=payload.name,
            description=payload.description,
            structure_json=payload.structure_json or {},
            is_default=payload.is_default,
        )
        session.add(tpl)
        await session.commit()
        await session.refresh(tpl)
        return ApiResponse(
            message="模板已创建",
            data={"id": tpl.id, "name": tpl.name, "is_default": tpl.is_default},
        )


@router.put("/lesson-templates/{template_id}")
async def update_lesson_template(template_id: int, payload: LessonTemplateUpdate):
    """更新教案模板"""
    async for session in get_session():
        tpl = await session.get(LessonTemplateORM, template_id)
        if not tpl:
            raise HTTPException(404, "模板不存在")
        if payload.name is not None:
            tpl.name = payload.name[:100] or tpl.name
        if payload.description is not None:
            tpl.description = payload.description
        if payload.structure_json is not None:
            tpl.structure_json = payload.structure_json
        if payload.is_default is not None and payload.is_default != tpl.is_default:
            if payload.is_default:
                r = await session.execute(select(LessonTemplateORM).where(LessonTemplateORM.is_default == True))
                for old in r.scalars().all():
                    old.is_default = False
                tpl.is_default = True
            else:
                # 取消当前默认 → 需要重新选一个全局默认
                if tpl.is_default:
                    tpl.is_default = False
                    r2 = await session.execute(
                        select(LessonTemplateORM)
                        .where(LessonTemplateORM.course_id.is_(None), LessonTemplateORM.id != template_id)
                        .order_by(LessonTemplateORM.created_at.asc())
                        .limit(1)
                    )
                    other = r2.scalars().first()
                    if other:
                        other.is_default = True
        await session.commit()
        return ApiResponse(message="模板已更新", data={"id": tpl.id, "name": tpl.name})


@router.delete("/lesson-templates/{template_id}")
async def delete_lesson_template(template_id: int):
    """删除教案模板（默认模板不允许删除）"""
    async for session in get_session():
        tpl = await session.get(LessonTemplateORM, template_id)
        if not tpl:
            raise HTTPException(404, "模板不存在")
        if tpl.is_default:
            return ApiResponse(
                success=False,
                message="默认模板不允许删除，请先将其他模板设为默认后重试",
                data={"error_code": "DEFAULT_TEMPLATE_PROTECTED", "fallbacks": ["set_other_default"]},
            )
        await session.delete(tpl)
        await session.commit()
        return ApiResponse(message="模板已删除", data={"id": template_id})


@router.post("/lesson-templates/{template_id}/set-default")
async def set_template_default(template_id: int):
    """设为默认模板"""
    async for session in get_session():
        tpl = await session.get(LessonTemplateORM, template_id)
        if not tpl:
            raise HTTPException(404, "模板不存在")
        r = await session.execute(select(LessonTemplateORM).where(LessonTemplateORM.is_default == True))
        for old in r.scalars().all():
            old.is_default = False
        tpl.is_default = True
        await session.commit()
        return ApiResponse(message="已设为默认模板", data={"id": template_id, "name": tpl.name})


@router.post("/lesson-templates/import")
async def import_lesson_template(file: UploadFile = File(...), course_id: Optional[int] = Form(None), name: Optional[str] = Form(None)):
    """上传Word文档(.docx)导入教案模板"""
    async for session in get_session():
        try:
            raw = await file.read()
            filename = (file.filename or "").lower()
            if filename.endswith(".docx"):
                structure = parse_template_docx(raw)
                template_name = name or (file.filename or "导入的模板").replace(".docx", "").replace(".DOCX", "")
                description = f"从Word文档导入: {file.filename or '未知'}"
            else:
                raise ValueError("仅支持 .docx 格式的Word文档，请上传可编辑的Word教案模板文件")
            tpl = LessonTemplateORM(
                course_id=course_id,
                name=str(template_name)[:100],
                description=str(description),
                structure_json=structure,
                is_default=False,
            )
            session.add(tpl)
            await session.commit()
            await session.refresh(tpl)
            return ApiResponse(message="模板导入成功", data={"id": tpl.id, "name": tpl.name})
        except Exception as e:
            return ApiResponse(
                success=False,
                message=f"模板导入失败: {type(e).__name__}: {e}",
                data={"error_code": "TEMPLATE_IMPORT_INVALID", "fallbacks": ["retry", "download_sample"]},
            )


@router.put("/lesson-templates/{template_id}/upload-docx")
async def upload_lesson_template_docx(template_id: int, file: UploadFile = File(...)):
    """上传Word文档(.docx)更新已有教案模板"""
    async for session in get_session():
        try:
            tpl = await session.get(LessonTemplateORM, template_id)
            if not tpl:
                raise HTTPException(404, "模板不存在")
            raw = await file.read()
            filename = (file.filename or "").lower()
            if not filename.endswith(".docx"):
                raise ValueError("仅支持 .docx 格式的Word文档")
            structure = parse_template_docx(raw)
            tpl.structure_json = structure
            await session.commit()
            return ApiResponse(message="模板已更新", data={"id": tpl.id, "name": tpl.name})
        except HTTPException:
            raise
        except Exception as e:
            return ApiResponse(
                success=False,
                message=f"模板更新失败: {type(e).__name__}: {e}",
                data={"error_code": "TEMPLATE_IMPORT_INVALID", "fallbacks": ["retry", "download_sample"]},
            )


@router.get("/lesson-templates/default/download")
async def download_default_template():
    """下载默认教案模板的Word文档(.docx)"""
    from pathlib import Path
    safe_name = urllib.parse.quote("默认教案模板")
    static_dir = Path(__file__).resolve().parent.parent / "static"
    default_path = static_dir / "default_template.docx"
    if default_path.exists():
        headers = {
            "Content-Disposition": f"attachment; filename*=UTF-8''{safe_name}.docx",
        }
        return Response(
            content=default_path.read_bytes(),
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            headers=headers,
        )
    async for session in get_session():
        r = await session.execute(
            select(LessonTemplateORM).where(LessonTemplateORM.is_default == True).limit(1)
        )
        tpl = r.scalars().first()
        if not tpl:
            raise HTTPException(404, "默认模板不存在")
        docx_bytes = generate_template_docx(tpl.structure_json or {}, "默认教案模板")
        headers = {
            "Content-Disposition": f"attachment; filename*=UTF-8''{safe_name}.docx",
        }
        return Response(
            content=docx_bytes,
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            headers=headers,
        )


@router.get("/lesson-templates/{template_id}/download")
async def download_lesson_template(template_id: int):
    """下载教案模板为可编辑的Word文档(.docx)"""
    async for session in get_session():
        tpl = await session.get(LessonTemplateORM, template_id)
        if not tpl:
            raise HTTPException(404, "模板不存在")
        docx_bytes = generate_template_docx(tpl.structure_json or {}, tpl.name or "教案模板")
        safe_name = re.sub(r'[\\/*?:"<>|]', '_', tpl.name or 'template')
        headers = {
            "Content-Disposition": f"attachment; filename*=UTF-8''{urllib.parse.quote(safe_name)}.docx",
        }
        return Response(
            content=docx_bytes,
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            headers=headers,
        )


@router.post("/lessons/{lesson_id}/fallback-template")
async def lesson_fallback_template(lesson_id: int, payload: dict):
    """跳过AI，直接用指定 template_id（或默认模板）生成骨架并写入 lesson.plan_json"""
    template_id = payload.get("template_id")
    async for session in get_session():
        lesson = await session.get(LessonORM, lesson_id)
        if not lesson:
            raise HTTPException(404, "教案不存在")
        course = await session.get(CourseORM, lesson.course_id)
        # 取模板
        template_structure = {}
        if template_id:
            tpl = await session.get(LessonTemplateORM, template_id)
            if tpl:
                template_structure = tpl.structure_json or {}
        if not template_structure:
            r = await session.execute(select(LessonTemplateORM).where(LessonTemplateORM.is_default == True).limit(1))
            dt = r.scalars().first()
            if dt:
                template_structure = dt.structure_json or {}
        defaults = template_structure.get("defaults", {}) if template_structure else {}
        plan_dict = await _apply_template_defaults(defaults, template_structure, course.name if course else "", lesson.chapter)
        lesson.plan_json = plan_dict
        lesson.updated_at = datetime.utcnow()
        await session.commit()
        return ApiResponse(message="已使用模板骨架填充", data={"id": lesson.id, "plan": plan_dict})


# ============================================================
# 一键提取知识点（多教材合并 + 联网校验 + 去重）
# ============================================================
async def _web_verify_knowledge(name: str, timeout: float = 3.0) -> tuple[int, str, str]:
    """联网校验知识点准确性：优先DuckDuckGo摘要/Wikipedia，失败降级warn；返回 (score 0-5, reason, flag)"""
    if not name:
        return (0, "知识点名称为空", "fail")
    name_enc = urllib.parse.quote(name)
    # 1) DuckDuckGo Instant Answer API
    try:
        import aiohttp
        async with aiohttp.ClientSession(timeout=aiohttp.ClientTimeout(total=timeout)) as http:
            async with http.get(f"https://api.duckduckgo.com/?q={name_enc}&format=json&no_html=1&skip_disambig=1") as resp:
                if resp.status == 200:
                    data = await resp.json(content_type=None)
                    abstract = (data.get("Abstract") or data.get("AbstractText") or "").strip()
                    if len(abstract) >= 15:
                        score = 5 if data.get("Type") == "A" else 4
                        reason_snippet = abstract[:140].replace("\n", " ")
                        return (score, f"DuckDuckGo摘要匹配: {reason_snippet}", "pass")
                    related = data.get("RelatedTopics") or []
                    if isinstance(related, list) and len(related) >= 1:
                        return (3, "DuckDuckGo找到相关主题，未找到精确摘要", "warn")
    except Exception:
        pass
    # 2) Wikipedia OpenSearch（中文）
    try:
        import aiohttp
        async with aiohttp.ClientSession(timeout=aiohttp.ClientTimeout(total=timeout)) as http:
            w_url = f"https://zh.wikipedia.org/w/api.php?action=opensearch&search={name_enc}&limit=2&format=json"
            async with http.get(w_url) as resp:
                if resp.status == 200:
                    j = await resp.json(content_type=None)
                    if isinstance(j, list) and len(j) >= 3 and j[2] and isinstance(j[2], list) and j[2][0]:
                        snippet = j[2][0][:140]
                        return (4, f"维基百科条目匹配: {snippet}", "pass")
    except Exception:
        pass
    # 3) 降级：LLM 自检（如果配置了）
    try:
        cfg = get_llm_config()
        if cfg.is_configured():
            llm = get_llm()
            text = await llm.chat(
                system_prompt="你是百科知识审核员。仅输出 JSON: {\"score\": 0-5整数, \"reason\": \"一句话\", \"flag\": \"pass|warn|fail\"}。score=0表示完全不存在/伪科学，3=边界/需要上下文，5=权威标准概念。",
                user_prompt=f"请评估知识点「{name}」是否真实存在且为主流知识。仅返回JSON，不要多余文字。",
                temperature=0.0,
                max_tokens=200,
            )
            # 尝试解析 JSON（可能是 code block）
            cleaned = text.strip()
            if cleaned.startswith("```"):
                cleaned = cleaned.strip("`")
                if cleaned.lower().startswith("json"):
                    cleaned = cleaned[4:]
                cleaned = cleaned.strip()
            start, end = cleaned.find("{"), cleaned.rfind("}")
            if start != -1 and end != -1:
                j = json.loads(cleaned[start:end+1])
                s = int(j.get("score", 2) or 2)
                f = j.get("flag", "warn") if j.get("flag") in ("pass","warn","fail") else "warn"
                r = str(j.get("reason", "LLM自检完成"))[:180]
                return (s, r, f)
    except Exception:
        pass
    return (2, "无法联网验证，仅保留AI原始识别结果", "warn")


@router.post("/courses/{course_id}/smart-extract-points")
async def smart_extract_points(course_id: int, payload: dict):
    """多教材批量提取知识点 + 去重 + 联网准确性校验 + 入库"""
    async for session in get_session():
        course = await session.get(CourseORM, course_id)
        if not course:
            raise HTTPException(404, "课程不存在")
        material_ids = payload.get("material_ids") or []
        if not isinstance(material_ids, list) or not material_ids:
            raise HTTPException(400, "请至少选择1个教材资源")
        # 读取教材内容并拼接
        mats_result = await session.execute(
            select(MaterialORM).where(
                MaterialORM.course_id == course_id,
                MaterialORM.id.in_([int(x) for x in material_ids]),
            )
        )
        mats = mats_result.scalars().all()
        if not mats:
            raise HTTPException(404, "所选教材不存在")
        combined_text_parts = []
        for m in mats:
            text = m.content_text or ""
            combined_text_parts.append(f"=== 教材：{m.filename} (类型:{MATERIAL_TYPE_LABELS.get(m.material_type,'其他')}) ===\n{text}")
        combined_text = "\n\n".join(combined_text_parts)

        # 使用分段递归提取（chunked_extract_knowledge），突破单次12000字符限制
        all_points: list[dict] = []
        try:
            chapter_label = "、".join(m.filename for m in mats)

            async def _progress_cb(current, total, points_so_far):
                await _update_extract_progress(course_id, current, total, points_so_far)

            kp_result = await chunked_extract_knowledge(
                course.name, chapter_label, combined_text,
                progress_callback=_progress_cb,
                subject=course.subject or None,
            )
            for p in kp_result.points:
                if isinstance(p, dict):
                    all_points.append(p)
                else:
                    try:
                        all_points.append(p.model_dump())
                    except Exception:
                        pass
        except Exception as e1:
            # 降级：调用 smart_extract（章节结构提取器），扁平化所有 knowledge_points
            try:
                # 降级链路也使用完整文本（不截断）
                chapters = await smart_extract(
                    course.name, "、".join(m.filename for m in mats), combined_text,
                    subject=course.subject or None,
                )
                def _flatten(nodes):
                    out = []
                    for node in (nodes or []):
                        for kp in (node.get("knowledge_points") or []):
                            out.append(kp)
                        out.extend(_flatten(node.get("children") or []))
                    return out
                all_points = _flatten(chapters)
            except Exception as e2:
                return ApiResponse(
                    success=False,
                    message=f"知识点提取均失败: 主链路{e1} / 降级链路{e2}",
                    data={"error_code": "EXTRACT_ALL_FAILED", "fallbacks": FALLBACK_ACTIONS},
                )

        # 归一化去重（精确匹配 + 模糊匹配）
        def _fuzzy_match_key(name: str, seen_keys: set[str]) -> str | None:
            lowered = name.lower().strip()
            if lowered in seen_keys:
                return lowered
            for k in seen_keys:
                if lowered in k or k in lowered:
                    return k
            for k in seen_keys:
                if SequenceMatcher(None, lowered, k).ratio() > 0.85:
                    return k
            return None

        seen: dict[str, dict] = {}
        seen_keys: set[str] = set()
        for p in all_points:
            nm = (p.get("name") or "").strip()
            if not nm:
                continue
            key = nm.lower()
            existing_key = _fuzzy_match_key(nm, seen_keys) if key not in seen else key
            if existing_key and existing_key in seen:
                old = seen[existing_key]
                old_def = old.get("definition", "")
                new_def = p.get("definition", "")
                if len(new_def) > len(old_def):
                    old["definition"] = new_def
                for flag in ("is_key_point", "is_difficult", "is_exam_point"):
                    if p.get(flag):
                        old[flag] = True
                old_imp = old.get("importance", 3)
                new_imp = p.get("importance", 3)
                old["importance"] = max(old_imp, new_imp)
                old_dif = old.get("difficulty", 3)
                new_dif = p.get("difficulty", 3)
                old["difficulty"] = max(old_dif, new_dif)
                old_prereq = set(old.get("prerequisites") or [])
                new_prereq = set(p.get("prerequisites") or [])
                old["prerequisites"] = list(old_prereq | new_prereq)
                old_rels = old.get("relationships") or []
                new_rels = p.get("relationships") or []
                existing_targets = {(r["target"], r["rel_type"]) for r in old_rels}
                for r in new_rels:
                    key_r = (r["target"], r["rel_type"])
                    if key_r not in existing_targets:
                        old_rels.append(r)
                        existing_targets.add(key_r)
                old["relationships"] = old_rels
            else:
                seen[key] = {
                    "name": nm,
                    "definition": (p.get("definition") or "").strip(),
                    "source_pages": (p.get("source_pages") or "").strip(),
                    "layer": p.get("layer") if p.get("layer") in ("basic","core","extension") else "core",
                    "importance": int(p.get("importance") or 3),
                    "difficulty": int(p.get("difficulty") or 3),
                    "is_key_point": bool(p.get("is_key_point", False)),
                    "is_difficult": bool(p.get("is_difficult", False)),
                    "is_exam_point": bool(p.get("is_exam_point", False)),
                    "prerequisites": list(p.get("prerequisites") or []),
                    "relationships": list(p.get("relationships") or []),
                }
                seen_keys.add(key)
        deduped = list(seen.values())

        # 质量过滤：剔除低质量/不完整知识点
        def _filter_quality(points: list[dict]) -> list[dict]:
            PLACEHOLDER_PATTERNS = ["待补充", "暂无", "待完善", "待填", "待定", "暂无定义", "暂无内容", "请补充", "请完善"]
            filtered = []
            seen_defs: list[str] = []
            for p in points:
                nm = (p.get("name") or "").strip()
                definition = (p.get("definition") or "").strip()
                if len(nm) < 2 or len(definition) < 15:
                    continue
                if any(pat in definition.lower() for pat in PLACEHOLDER_PATTERNS):
                    continue
                importance = int(p.get("importance") or 3)
                difficulty = int(p.get("difficulty") or 3)
                if not (1 <= importance <= 5) or not (1 <= difficulty <= 5):
                    continue
                is_dup = False
                for ed in seen_defs:
                    if SequenceMatcher(None, definition, ed).ratio() > 0.9:
                        is_dup = True
                        break
                if is_dup:
                    continue
                seen_defs.append(definition)
                filtered.append(p)
            return filtered

        pre_count = len(deduped)
        deduped = _filter_quality(deduped)
        if len(deduped) < pre_count:
            logger.info("质量过滤移除了 %d 个低质量知识点", pre_count - len(deduped))

        # 联网校验（批量：>30 个知识点分批，每批单例，总数限制 200）
        limited = deduped[:200]
        accuracy_results = []
        for kp in limited:
            score, reason, flag = await _web_verify_knowledge(kp["name"])
            kp["accuracy"] = {"accuracy_score": score, "accuracy_reason": reason, "accuracy_flag": flag}
            accuracy_results.append({"name": kp["name"], "score": score, "flag": flag})

        # 写 DB（先删除课程原有 KP？不，追加但重名更新。这里简单：追加，按 name+course_id 判定若存在则 update）
        written_count = 0
        output_points = []
        for kp in limited:
            # 查重同课程同名
            exist_r = await session.execute(
                select(KnowledgePointORM).where(
                    KnowledgePointORM.course_id == course_id,
                    func.lower(KnowledgePointORM.name) == kp["name"].lower(),
                ).limit(1)
            )
            exist = exist_r.scalars().first()
            accuracy = kp.get("accuracy") or {}
            if exist:
                exist.definition = kp["definition"] or exist.definition
                exist.source_pages = kp["source_pages"] or exist.source_pages
                exist.layer = kp["layer"]
                exist.importance = kp["importance"]
                exist.difficulty = kp["difficulty"]
                exist.is_key_point = int(kp["is_key_point"])
                exist.is_difficult = int(kp["is_difficult"])
                exist.is_exam_point = int(kp["is_exam_point"])
                exist.prerequisites_json = kp["prerequisites"]
                exist.relationships_json = kp.get("relationships") or []
                exist.accuracy_json = accuracy
                kp_id = exist.id
            else:
                obj = KnowledgePointORM(
                    course_id=course_id,
                    chapter="",
                    name=kp["name"],
                    definition=kp["definition"],
                    source_pages=kp["source_pages"],
                    layer=kp["layer"],
                    importance=kp["importance"],
                    difficulty=kp["difficulty"],
                    is_key_point=int(kp["is_key_point"]),
                    is_difficult=int(kp["is_difficult"]),
                    is_exam_point=int(kp["is_exam_point"]),
                    prerequisites_json=kp["prerequisites"],
                    relationships_json=kp.get("relationships") or [],
                    accuracy_json=accuracy,
                    sort_order=written_count,
                )
                session.add(obj)
                await session.flush()
                kp_id = obj.id
            written_count += 1
            output_points.append({"id": kp_id, **kp})
        await session.commit()

        # ---- 自动生成章节目录 ----
        chapter_tree = []
        try:
            raw_chapters = await organize_knowledge_into_chapters(course.name, deduped)
            if raw_chapters:
                # 删除该课程下原有的自动生成章节（非手动创建的，即没有 parent_id 的顶级章节）
                existing_chapters_r = await session.execute(
                    select(ChapterORM).where(
                        ChapterORM.course_id == course_id,
                        ChapterORM.parent_id.is_(None),
                    )
                )
                for old_ch in existing_chapters_r.scalars().all():
                    await session.delete(old_ch)
                await session.flush()

                for ci, ch in enumerate(raw_chapters):
                    ch_name = ch.get("name", "").strip()
                    if not ch_name:
                        continue
                    ch_obj = ChapterORM(
                        course_id=course_id,
                        parent_id=None,
                        name=ch_name,
                        sort_order=ci,
                    )
                    session.add(ch_obj)
                    await session.flush()
                    children = ch.get("children") or []
                    for si, child in enumerate(children):
                        child_name = child.get("name", "").strip()
                        if not child_name:
                            continue
                        child_obj = ChapterORM(
                            course_id=course_id,
                            parent_id=ch_obj.id,
                            name=child_name,
                            sort_order=si,
                        )
                        session.add(child_obj)
                        await session.flush()
                        kp_names = child.get("knowledge_points") or []
                        for kpn in kp_names:
                            kp_r = await session.execute(
                                select(KnowledgePointORM).where(
                                    KnowledgePointORM.course_id == course_id,
                                    func.lower(KnowledgePointORM.name) == kpn.strip().lower(),
                                ).limit(1)
                            )
                            kp = kp_r.scalars().first()
                            if kp:
                                kp.chapter_id = child_obj.id
                                kp.chapter = child_name
                await session.commit()
                chapter_tree = raw_chapters
        except Exception as e:
            # 章节生成失败不阻塞主流程，仅记录
            pass

        stats = {
            "total": len(output_points),
            "pass": sum(1 for a in accuracy_results if a["flag"] == "pass"),
            "warn": sum(1 for a in accuracy_results if a["flag"] == "warn"),
            "fail": sum(1 for a in accuracy_results if a["flag"] == "fail"),
            "avg_score": round(sum(a["score"] for a in accuracy_results) / max(1, len(accuracy_results)), 2),
        }
        return ApiResponse(
            message=f"一键提取完成：共识别 {len(output_points)} 个知识点（通过{stats['pass']} / 待确认{stats['warn']} / 不通过{stats['fail']}）",
            data={
                "course_id": course_id,
                "points": output_points,
                "accuracy_stats": stats,
                "material_count": len(mats),
                "chapters": chapter_tree,
            },
        )


# ============================================================
# 知识点 XLSX 导出（贴合用户模板：Sheet3 R1说明 R2表头 14列）
# ============================================================
def _layer_to_category(layer: str, is_category_node: bool = False) -> str:
    if is_category_node:
        return "元认知"
    return {
        "basic": "事实性",
        "core": "概念性",
        "extension": "程序性",
    }.get(layer, "概念性")


async def _build_hierarchy(session: AsyncSession, course_id: int, filter_ids: list[int] | None):
    """从 DB KP 数据构造层级路径：取 layer 分桶 + 按 prerequisites 推断后置/关联；输出扁平记录列表（每节点一行）"""
    stmt = select(KnowledgePointORM).where(KnowledgePointORM.course_id == course_id)
    if filter_ids:
        stmt = stmt.where(KnowledgePointORM.id.in_(filter_ids))
    stmt = stmt.order_by(KnowledgePointORM.layer, KnowledgePointORM.sort_order)
    r = await session.execute(stmt)
    kps = list(r.scalars().all())
    # 分层：basic(事实性) 放 B 列父节点，core(概念性) 作为子节点 C 列，extension(程序性) D 列；余下列留空（最多 7 层 B-H）
    # 实际策略：用户模板 B-H=节点名称(层级路径)。这里采用 前缀层：课程(固定)->章(layer桶)->节(name prefix) 共3层，然后知识点名称放后面列
    layers = {"basic": [], "core": [], "extension": []}
    for kp in kps:
        bucket = kp.layer if kp.layer in layers else "core"
        layers[bucket].append(kp)
    # 反向构建 后置节点图
    prereq_map: dict[str, list[str]] = {}
    for kp in kps:
        name = kp.name
        for pr in (kp.prerequisites_json or []):
            if isinstance(pr, str) and pr:
                prereq_map.setdefault(pr, []).append(name)
    records = []
    # 课程分类节点（行1：课程名称根节点）
    course_r = await session.get(CourseORM, course_id)
    course_name = course_r.name if course_r else "课程"
    records.append({
        "node_type": "分类",
        "path": [course_name, "", "", "", "", "", ""],  # B-H
        "pre": "", "post": "", "rel": "",
        "tags": "",
        "category": "元认知",
        "definition": f"课程根节点：{course_name}，共{len(kps)}个知识点",
    })
    # 三个 layer 桶作为"章"分类节点
    layer_label = {"basic": "基础知识点（事实性）", "core": "核心知识点（概念性）", "extension": "拓展知识点（程序性）"}
    for layer_key in ("basic", "core", "extension"):
        bucket = layers[layer_key]
        if not bucket:
            continue
        records.append({
            "node_type": "分类",
            "path": [course_name, layer_label[layer_key], "", "", "", "", ""],
            "pre": "", "post": "", "rel": "",
            "tags": "",
            "category": "元认知",
            "definition": f"{layer_label[layer_key]}，共{len(bucket)}个知识点",
        })
        # 每个 KP 作为 知识点节点
        for kp in bucket:
            tags_list = []
            if kp.is_key_point:
                tags_list.append("重点")
            if kp.is_difficult:
                tags_list.append("难点")
            if kp.is_exam_point:
                tags_list.append("考点")
            acc = kp.accuracy_json or {}
            if isinstance(acc, dict) and acc.get("accuracy_flag") == "fail":
                tags_list.append("待核实")
            # path：课程 → 章桶 → 知识点名（其余列空，总层数<7）
            path = [course_name, layer_label[layer_key], kp.name, "", "", "", ""]
            pre_list = [p for p in (kp.prerequisites_json or []) if isinstance(p, str) and p]
            post_list = prereq_map.get(kp.name, [])
            records.append({
                "node_type": "知识点",
                "path": path,
                "pre": "；".join(pre_list),
                "post": "；".join(post_list),
                "rel": "",
                "tags": "；".join(tags_list),
                "category": _layer_to_category(layer_key),
                "definition": kp.definition or "（暂无定义）",
            })
    return records


@router.post("/courses/{course_id}/knowledge-points/export-xlsx")
async def export_knowledge_xlsx(course_id: int, payload: dict | None = None):
    """按模板(Sheet3, R1说明, R2表头)导出 14 列 XLSX"""
    payload = payload or {}
    filter_ids = payload.get("filter_ids") or None
    chapter_name = payload.get("chapter_name") or ""
    # 模板原第一行/第二行说明文字（完全贴合用户模板格式）
    TEMPLATE_R1 = "说明：1)节点类型必填：分类/知识点；2)列B-H填节点层级路径，分类节点先填，知识点节点逐层往后填直至首个空列；3)前置/后置/关联节点填名称并以分号分隔；4)标签填重点/难点/考点/思政等；5)知识点分类∈{事实性,概念性,程序性,元认知}"
    TEMPLATE_HEADER = [
        "节点类型*",                # A
        "节点名称",                  # B
        "节点名称",                  # C
        "节点名称",                  # D
        "节点名称",                  # E
        "节点名称",                  # F
        "节点名称",                  # G
        "节点名称",                  # H
        "前置节点",                  # I
        "后置节点",                  # J
        "关联节点",                  # K
        "标签",                      # L
        "知识点分类",                # M
        "节点说明",                  # N
    ]
    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(500, "缺少openpyxl依赖，请在backend目录执行 pip install openpyxl")

    async for session in get_session():
        course = await session.get(CourseORM, course_id)
        if not course:
            raise HTTPException(404, "课程不存在")
        records = await _build_hierarchy(session, course_id, filter_ids)

    wb = Workbook()
    # 删除默认 sheet，创建名为 Sheet3 的工作表（与用户模板一致）
    default = wb.active
    wb.remove(default)
    ws = wb.create_sheet("Sheet3")
    # R1 说明（A列整长文字，其余空）
    ws.append([TEMPLATE_R1] + [""] * 13)
    # R2 表头（14列）
    ws.append(TEMPLATE_HEADER)
    # R3 起数据
    for rec in records:
        path = rec["path"]  # list 7 项 (B-H)
        row = [
            rec["node_type"],          # A
            path[0], path[1], path[2], path[3], path[4], path[5], path[6],  # B-H
            rec["pre"],                # I
            rec["post"],               # J
            rec["rel"],                # K
            rec["tags"],               # L
            rec["category"],           # M
            rec["definition"],         # N
        ]
        ws.append(row)
    # 列宽优化（避免过窄）
    widths = [10, 22, 22, 22, 22, 22, 22, 22, 22, 22, 16, 22, 12, 60]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx) if idx <= 26 else "A" + chr(64 + idx - 26)].width = w
    # 标题行 R1 合并 A1:N1，保持第一行作为说明
    from openpyxl.styles import Alignment, Font
    ws.merge_cells("A1:N1")
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["A1"].font = Font(italic=True, color="666666")

    # 转 bytes
    import io
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    safe_title = f"{course.name}_知识点图谱".replace(" ", "_")
    safe_title = "".join(c for c in safe_title if c.isalnum() or c in "_-") or "knowledge_points"
    filename_encoded = urllib.parse.quote(f"{safe_title}.xlsx")
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=\"knowledge_points.xlsx\"; filename*=UTF-8''{filename_encoded}",
        },
    )


# ============================================================
# 知识点提取进度轮询
# ============================================================
@router.get("/courses/{course_id}/extract-progress")
async def get_extract_progress(course_id: int):
    progress = _extract_progress.get(course_id)
    if not progress:
        return ApiResponse(data={
            "current": 0,
            "total": 0,
            "points": [],
            "updated_at": None,
            "running": False,
        })
    return ApiResponse(data={
        "current": progress["current"],
        "total": progress["total"],
        "points": progress["points"],
        "updated_at": progress["updated_at"],
        "running": progress["current"] < progress["total"],
    })
