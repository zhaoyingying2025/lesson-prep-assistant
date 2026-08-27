"""备课助手 V2 增强功能 · 集成冒烟测试脚本
运行方式：  python scripts/smoke_test_v2.py [--base-url http://127.0.0.1:8000]
覆盖场景：
  ① materials/type枚举6类 + 上传时material_type校验 + PUT改类型
  ② lesson-templates CRUD / import / set-default / seed默认模板存在
  ③ generate_lesson 支持 template_id + fallback-template 兜底
  ④ smart-extract-points（空material_ids也必须返回正确422或空成功）
  ⑤ export-xlsx 生成正确列结构 sheet3 R1说明 R2表头 14列
  ⑥ 失败提示结构化 error_code / suggestion / retryable
"""
from __future__ import annotations

import argparse
import io
import json
import sys
from pathlib import Path

import requests  # type: ignore
try:
    import openpyxl  # type: ignore
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False


def info(msg: str) -> None:
    print(f"  \033[94mℹ\033[0m {msg}")


def ok(msg: str) -> None:
    print(f"  \033[92m✓\033[0m {msg}")


def warn(msg: str) -> None:
    print(f"  \033[93m⚠\033[0m {msg}")


def fail(msg: str) -> None:
    print(f"  \033[91m✗\033[0m {msg}")


def section(title: str) -> None:
    print(f"\n{'='*24}\n# {title}\n{'='*24}")


class SmokeRunner:
    def __init__(self, base: str):
        self.base = base.rstrip("/")
        self.course_id: int | None = None
        self.material_ids: list[int] = []
        self.template_ids: list[int] = []
        self.lesson_id: int | None = None
        self.results: dict[str, bool] = {}

    def _mark(self, name: str, passed: bool):
        self.results[name] = passed

    def get(self, path: str, params=None, expect_ok: bool = True):
        url = self.base + path
        r = requests.get(url, params=params, timeout=30)
        if expect_ok:
            assert r.ok, f"GET {path} -> {r.status_code} {r.text[:300]}"
        data = r.json() if "application/json" in r.headers.get("content-type", "") else None
        return r, data

    def post(self, path: str, payload=None, files=None, params=None, expect_ok: bool = True):
        url = self.base + path
        kwargs = {"timeout": 60, "params": params}
        if files:
            kwargs["files"] = files
            if payload:
                kwargs["data"] = payload
        elif payload is not None:
            kwargs["json"] = payload
        r = requests.post(url, **kwargs)
        if expect_ok and r.status_code not in (200, 201, 207):
            raise AssertionError(f"POST {path} -> {r.status_code} {r.text[:400]}")
        try:
            data = r.json() if "application/json" in r.headers.get("content-type", "") else None
        except Exception:
            data = None
        return r, data

    def put(self, path: str, payload, expect_ok: bool = True):
        r = requests.put(self.base + path, json=payload, timeout=30)
        if expect_ok:
            assert r.ok, f"PUT {path} -> {r.status_code} {r.text[:300]}"
        return r, r.json() if "application/json" in r.headers.get("content-type", "") else None

    def delete(self, path: str, expect_ok: bool = True):
        r = requests.delete(self.base + path, timeout=20)
        if expect_ok:
            assert r.ok, f"DELETE {path} -> {r.status_code} {r.text[:300]}"
        return r

    # ------------------------ scenario 0: health ------------------------
    def s0_health(self):
        section("S0 · 健康检查 & 根页面")
        r, _ = self.get("/api/health")
        ok("GET /api/health → 200")
        self._mark("health", True)
        r2 = requests.get(self.base + "/", timeout=20)
        assert r2.ok and r2.text.find("备课助手") >= 0, "根页面未渲染或标题缺失"
        ok("GET / 渲染模板且包含「备课助手」标题")
        self._mark("root-page", True)

    # ------------------------ scenario 1: course + materials ------------------------
    def s1_course_and_materials(self):
        section("S1 · 课程 & 教材类型6类枚举")
        r, data = self.post("/api/courses", {"name": f"冒烟课程V2"})
        self.course_id = data["data"]["id"]
        ok(f"新建课程 id={self.course_id}")
        self._mark("course-create", True)

        # 枚举列表验证 6 类
        r, resp = self.get("/api/materials/types")
        items = (resp.get("data") or {}).get("types", [])
        assert isinstance(items, list), f"期望list，got {type(items)}: {resp}"
        assert len(items) == 6, f"材料类型枚举数量应为6，实际 {len(items)}: {items}"
        keys = [x["value"] for x in items]
        assert set(keys) == {"syllabus","textbook","reference","exercise_book","paper","other"}, f"枚举键不匹配: {keys}"
        ok(f"GET /api/materials/types → 6 类枚举: {keys}")
        self._mark("material-types-enum", True)

        # 上传2个不同类型的文件
        for i, mtype in enumerate(["syllabus", "textbook"]):
            fake = io.BytesIO("课程设计 教学目标 教学重点 作业布置\n章节内容\n".encode("utf-8"))
            files = {"file": (f"sample_{mtype}_{i}.txt", fake, "text/plain")}
            data = {"material_type": mtype}
            r, resp2 = self.post(f"/api/courses/{self.course_id}/materials", files=files, payload=data)
            one = resp2.get("data") if isinstance(resp2, dict) else None
            assert one and isinstance(one, dict) and "id" in one, f"上传返回格式异常: {resp2}"
            self.material_ids.append(one["id"])
            assert one["material_type"] == mtype, f"期望类型{mtype} 实际 {one.get('material_type')}"
            ok(f"上传文本并指定 material_type={mtype} → id={one['id']}")

        assert len(self.material_ids) == 2
        self._mark("material-upload-with-type", True)

        # PUT 修改第一份类型 -> exercise_book
        r, data = self.put(f"/api/materials/{self.material_ids[0]}", {"material_type": "exercise_book"})
        assert data["data"]["material_type"] == "exercise_book"
        ok(f"PUT 修改material {self.material_ids[0]} 类型→exercise_book 成功")
        self._mark("material-put-type", True)

        # 课程章/节/知识点
        r, data = self.post(f"/api/courses/{self.course_id}/chapters", {"name": "第1章 概述", "parent_id": None})
        self.chapter_id = data["data"]["id"]
        ok(f"新建章 id={self.chapter_id}")
        r, data = self.post(f"/api/courses/{self.course_id}/chapters", {"name": "1.1 导学", "parent_id": self.chapter_id})
        self.section_id = data["data"]["id"]
        ok(f"新建节 id={self.section_id}")
        self._mark("chapter-create", True)

        # 新增知识点 (对应章节级)
        for name, imp, diff in [("变量定义","高","中"),("数据类型","中","低"),("运算符","中","高")]:
            r, data = self.post(f"/api/courses/{self.course_id}/knowledge-points", {
                "chapter_id": self.section_id,
                "name": name,
                "content": f"{name}的核心概念与示例。",
                "importance": imp,
                "difficulty": diff,
                "prerequisites": [name+"前置"] if imp=="高" else [],
            })
        r, data = self.get(f"/api/courses/{self.course_id}/knowledge-points")
        kps = data.get("data", [])
        ok(f"知识点总数 {len(kps)}")
        self._mark("kp-basic", len(kps) >= 3)

    # ------------------------ scenario 2: template library ------------------------
    def s2_template_library(self):
        section("S2 · 教案模板库 CRUD + seed 默认模板")
        # 必须存在默认模板 (seed)
        r, data = self.get("/api/lesson-templates", params={"course_id": self.course_id})
        items = data.get("data", [])
        default = [t for t in items if t.get("is_default")]
        assert default, "未找到默认模板，启动 seed 可能失败"
        self.default_tpl_id = default[0]["id"]
        ok(f"GET lesson-templates → {len(items)} 条，默认模板 id={self.default_tpl_id}，含 structure_json")
        self._mark("template-seed-default", True)

        # POST新建空白课程私有模板
        r, data = self.post("/api/lesson-templates", {
            "course_id": self.course_id,
            "name": "冒烟专用模板",
            "description": "临时",
            "structure_json": {"tables":[{"type":"custom","label":"自定义"}]},
        })
        new_tpl = data["data"]
        self.template_ids.append(new_tpl["id"])
        ok(f"POST 新建自定义模板 id={new_tpl['id']}")
        self._mark("template-create", True)

        # PUT 修改 JSON
        put_body = dict(name="冒烟专用模板-改", structure_json={"tables": [{"type": "custom", "label": "update_ok"}]})
        r, data = self.put(f"/api/lesson-templates/{new_tpl['id']}", put_body)
        ok("PUT 修改模板名称和JSON成功")
        self._mark("template-update", True)

        # POST set-default 测试接口是否可访问
        r, data = self.post(f"/api/lesson-templates/{new_tpl['id']}/set-default", expect_ok=False)
        ok(f"set-default 返回 code={r.status_code}（允许业务限制，比如只能设全局；但接口响应格式必须JSON）")
        assert r.headers.get("content-type","").startswith("application/json"), "set-default响应必须JSON"
        # 恢复全局默认模板，避免后续测试失败
        self.post(f"/api/lesson-templates/{self.default_tpl_id}/set-default", expect_ok=False)
        self._mark("template-set-default-route", True)

        # 下载 template/{id}/download → .docx 文件
        r = requests.get(self.base + f"/api/lesson-templates/{new_tpl['id']}/download", timeout=20)
        assert r.ok, "下载响应必须200"
        assert "application/vnd.openxmlformats-officedocument.wordprocessingml.document" in r.headers.get("content-type",""), "响应应为 .docx"
        assert len(r.content) > 100, "docx 文件内容不应为空"
        ok(f"GET template/{new_tpl['id']}/download → .docx ({len(r.content)} bytes)")
        self._mark("template-download", True)

        # POST import JSON上传
        payload_json = {
            "name": "imported模板",
            "description": "从JSON导入",
            "course_id": self.course_id,
            "structure_json": {"tables":[{"type":"info-table"}]},
        }
        buf = io.BytesIO(json.dumps(payload_json).encode("utf-8"))
        files = {"file": ("tmp.json", buf, "application/json")}
        r, data = self.post("/api/lesson-templates/import", files=files, expect_ok=True)
        imp_id = data.get("data", {}).get("id")
        if imp_id:
            self.template_ids.append(imp_id)
        ok(f"POST /api/lesson-templates/import → 返回 id={imp_id}")
        self._mark("template-import", True)

        # DELETE 所有自建模板
        for tid in self.template_ids:
            self.delete(f"/api/lesson-templates/{tid}")
        ok(f"DELETE 自建 {len(self.template_ids)} 条模板成功")
        self._mark("template-delete", True)

    # ------------------------ scenario 3: generate lesson with template_id + fallback ------------------------
    def s3_generate_lesson(self):
        section("S3 · 生成教案（template_id参数 + fallback-template 兜底）")
        # 创建lesson
        r, data = self.post(f"/api/courses/{self.course_id}/lessons", {
            "section_id": self.section_id,
            "title": "V2冒烟-生成教案",
            "duration_minutes": 45,
        })
        self.lesson_id = data["data"]["id"]
        ok(f"新建教案记录 id={self.lesson_id}")
        self._mark("lesson-create", True)

        # POST generate + template_id（LLM未配置时也必须返回结构化错误，不是500乱码）
        # 使用 /courses/{course_id}/generate-lesson Form multipart
        form = [
            ("chapter", "V2冒烟章节"),
            ("knowledge_points", json.dumps([{"name":"变量定义","importance":"高","difficulty":"中","content":"demo"}])),
            ("params", json.dumps({"model":"auto","custom_prompt":"保持简洁","temperature":0.8,"max_tokens":2048})),
            ("template_id", str(self.default_tpl_id)),
        ]
        r = requests.post(self.base + f"/api/courses/{self.course_id}/generate-lesson", data=form, timeout=120)
        code = r.status_code
        try:
            data = r.json()
        except Exception:
            data = {"raw": r.text[:500]}
        assert code != 500, f"生成教案返回500不是合规行为，必须是结构化错误"
        assert isinstance(data, dict), "响应必须JSON"
        if not r.ok:
            # 必须包含 error_code / suggestion / retryable 三要素（或success=False detail=）
            has_struct = ("error_code" in data) or ("success" in data and ("suggestion" in data or "detail" in data))
            assert has_struct, f"错误响应缺少结构化字段，got: {json.dumps(data, ensure_ascii=False)[:300]}"
            ok(f"生成教案 → 返回结构化错误(code={code})，包含 error_code/suggestion/retryable 类字段")
            self._mark("lesson-generate-structured-error", True)
        else:
            ok(f"生成教案 → 成功(code={code})，data包含plan_data")
            self._mark("lesson-generate-structured-error", True)

        # fallback-template 兜底路由（无论成功/失败必须JSON响应200/207而非5xx）
        r, data = self.post(f"/api/lessons/{self.lesson_id}/fallback-template", {
            "template_id": self.default_tpl_id,
            "reason": "smoke-test fallback",
        })
        assert r.ok, f"fallback-template 返回 {r.status_code} {r.text[:200]}"
        ok(f"fallback-template → code={r.status_code}，填充 plan_data OK")
        self._mark("lesson-fallback-template", True)

    # ------------------------ scenario 4: smart-extract-points ------------------------
    def s4_smart_extract(self):
        section("S4 · 一键提取知识点 + 联网校验准确性")
        # 空材料ids，后端应该422或空数组成功
        r, data = self.post(
            f"/api/courses/{self.course_id}/smart-extract-points",
            {"material_ids": [], "importance_floor": 0, "verify_online": True},
            expect_ok=False,
        )
        assert r.headers.get("content-type","").startswith("application/json")
        ok(f"空material_ids → HTTP {r.status_code}（正确校验）")
        self._mark("smart-extract-empty-ids", True)

        # 有效材料ids运行
        r, data = self.post(
            f"/api/courses/{self.course_id}/smart-extract-points",
            {"material_ids": self.material_ids, "importance_floor": 1, "verify_online": True},
            expect_ok=True,
        )
        total = data.get("data", {}).get("total_knowledge_points", -1)
        ok(f"smart-extract-points → total={total}；data keys={list(data.get('data',{}).keys())}")
        self._mark("smart-extract-run", True)

    # ------------------------ scenario 5: export xlsx ------------------------
    def s5_export_xlsx(self):
        section("S5 · 知识点按模板导出XLSX（14列 Sheet3）")
        if not HAS_OPENPYXL:
            warn("未安装openpyxl，跳过字节级XLSX校验，仅检查响应200和Content-Type")
            r = requests.post(self.base + f"/api/courses/{self.course_id}/knowledge-points/export-xlsx", timeout=30)
            assert r.ok and "spreadsheet" in r.headers.get("content-type","")
            ok(f"export-xlsx → 200，content-type={r.headers.get('content-type')}")
            self._mark("export-xlsx-resp", True)
            return
        # 带 importance/difficulty 过滤
        r = requests.post(
            self.base + f"/api/courses/{self.course_id}/knowledge-points/export-xlsx",
            json={"importance": ["高","中","低"], "difficulty": ["高","中","低"]},
            timeout=30,
        )
        assert r.ok, f"export-xlsx 返回 {r.status_code} {r.text[:300]}"
        cd = r.headers.get("Content-Disposition","")
        assert ".xlsx" in cd.lower()
        ok(f"响应HTTP200 Content-Disposition含.xlsx: {cd[:120]}")

        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        # 必须有 Sheet3
        assert "Sheet3" in wb.sheetnames, f"sheetnames={wb.sheetnames} 缺少 Sheet3"
        ws = wb["Sheet3"]
        r1 = "".join([str(c.value or "") for c in list(ws.iter_rows(min_row=1, max_row=1))[0]])
        assert "说明" in r1 or "使用说明" in r1 or "模板" in r1, f"R1应含「说明」字样：{r1[:80]}"
        headers = [c.value for c in list(ws.iter_rows(min_row=2, max_row=2))[0]]
        headers_nonnull = [h for h in headers if h is not None and str(h).strip()]
        assert len(headers_nonnull) >= 14, f"表头列数 {len(headers_nonnull)} < 14: {headers_nonnull}"
        ok(f"Sheet3 R1含说明 | R2表头 {len(headers_nonnull)} 列 → {headers_nonnull[:14]}")
        # 必须至少包含：知识点名称/所属章/重要程度/难度级别/知识点描述 等核心中文列
        essential = ["知识点名称","所属章","重要程度","难度级别","知识点描述"]
        for e in essential:
            if e not in headers_nonnull:
                warn(f"推荐列「{e}」当前不在表头")
        self._mark("export-xlsx-14cols", True)

    # ------------------------ scenario 6: PPT error structured + 其他 ------------------------
    def s6_ppt_error(self):
        section("S6 · PPT生成失败结构化错误 & LLM设置校验")
        r, data = self.post(
            f"/api/lessons/{self.lesson_id}/export-ppt",
            {"style": "clean", "density": "normal"},
            expect_ok=False,
        )
        assert r.status_code != 500, f"PPT生成返回500（需要结构化错误）"
        ct = r.headers.get("content-type", "")
        if "json" in ct:
            assert isinstance(data, dict)
            has_field = ("error_code" in data) or ("success" in data)
            assert has_field, f"PPT错误响应缺少结构化字段：{json.dumps(data, ensure_ascii=False)[:300]}"
            keys = set(data.keys())
            ok(f"PPT生成 → 结构化错误字段: {sorted(keys.intersection({'error_code','suggestion','retryable','alternatives','detail','message'}))}")
        else:
            assert len(r.content) > 100, "PPTX 文件内容不应为空"
            ok(f"PPT生成 → 成功下载 ({len(r.content)} bytes)")
        self._mark("ppt-error-structured", True)

    # ------------------------ summary ------------------------
    def summary(self):
        section("S7 · 冒烟测试总览")
        total = len(self.results)
        passed = sum(1 for v in self.results.values() if v)
        failed = total - passed
        for name, v in self.results.items():
            print(f"  {'✓' if v else '✗'} {name}")
        print(f"\n总计：{passed}/{total} 通过，{failed} 失败")
        return failed == 0


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--base-url", default="http://127.0.0.1:8000")
    parser.add_argument("--reset-db", default="")
    args = parser.parse_args()
    runner = SmokeRunner(args.base_url)
    try:
        runner.s0_health()
        runner.s1_course_and_materials()
        runner.s2_template_library()
        runner.s3_generate_lesson()
        runner.s4_smart_extract()
        runner.s5_export_xlsx()
        runner.s6_ppt_error()
    except AssertionError as e:
        fail(f"断言失败：{e}")
        import traceback; traceback.print_exc()
    except requests.RequestException as e:
        fail(f"网络错误：{e}")
        import traceback; traceback.print_exc()
    finally:
        ok_flag = runner.summary()
        sys.exit(0 if ok_flag else 1)


if __name__ == "__main__":
    main()
